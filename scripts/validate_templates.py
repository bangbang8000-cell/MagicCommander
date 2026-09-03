"""MagicCommander 模板库健康校验（4.0.0-F0-1，对齐 AL scripts/validate_templates.py）

校验 example/ 模板库（模板中心数据源）每个模板项目：
  - 目录结构：templates/ 存在且含 ≥1 个 .j2；para.xlsx 或 excel/ 存在
  - template.meta.json 存在、可解析，必填字段齐全（name/description/scenario/inputRequirements/outputDescription）
  - 每个 .j2 可被 Jinja2 编译（语法健康）
  - excel/*.xlsx 可被 openpyxl 读取（数据健康）
  - 渲染健康：临时 workspace dry-run 渲染成功且产出 ≥1 台设备、无 [渲染错误]

注意：后端 config.py 在导入时读取 MC_WORKSPACE，故使用单一临时 workspace 一次性渲染全部模板。

用法：
  python scripts/validate_templates.py
"""
import json
import logging
import os
import re
import shutil
import sys
import tempfile

REPO = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
BACKEND = os.path.join(REPO, 'backend')
base = os.path.join(REPO, 'example')

# template.meta.json 必填字段
REQUIRED_META_FIELDS = ['name', 'description', 'scenario', 'inputRequirements', 'outputDescription']


def _discover_templates():
    if not os.path.isdir(base):
        return []
    return sorted(
        name
        for name in os.listdir(base)
        if os.path.isdir(os.path.join(base, name))
        and not name.startswith(('.', '_'))
        and (
            os.path.isdir(os.path.join(base, name, 'templates'))
            or os.path.exists(os.path.join(base, name, 'para.xlsx'))
        )
    )


def _check_structure(tpl_dir, problems):
    """1. 目录结构检查"""
    templates_dir = os.path.join(tpl_dir, 'templates')
    if not os.path.isdir(templates_dir):
        problems.append('缺少 templates/ 目录')
        return None
    j2_files = sorted(
        f for f in os.listdir(templates_dir)
        if f.endswith('.j2') or f.endswith('.jinja') or f.endswith('.jinja2')
    )
    if not j2_files:
        problems.append('templates/ 下无 .j2 模板')
        return None
    has_para = os.path.exists(os.path.join(tpl_dir, 'para.xlsx'))
    excel_dir = os.path.join(tpl_dir, 'excel')
    has_excel = os.path.isdir(excel_dir) and any(
        f.lower().endswith(('.xlsx', '.xls')) for f in os.listdir(excel_dir)
    )
    if not has_para and not has_excel:
        problems.append('缺少 para.xlsx 且 excel/ 目录无数据文件（无法渲染）')
    return j2_files


def _check_meta(tpl_dir, problems):
    """2. template.meta.json 完整性"""
    meta_path = os.path.join(tpl_dir, 'template.meta.json')
    if not os.path.exists(meta_path):
        problems.append('缺少 template.meta.json（模板中心元数据）')
        return
    try:
        with open(meta_path, encoding='utf-8') as f:
            meta = json.load(f)
    except (OSError, ValueError) as e:
        problems.append(f'template.meta.json 解析失败: {e}')
        return
    missing = [k for k in REQUIRED_META_FIELDS if not meta.get(k)]
    if missing:
        problems.append(f'template.meta.json 缺少必填字段: {missing}')


def _check_jinja(tpl_dir, j2_files, problems):
    """3. Jinja2 语法健康"""
    from jinja2 import Environment, FileSystemLoader, StrictUndefined
    env = Environment(
        loader=FileSystemLoader(os.path.join(tpl_dir, 'templates')),
        undefined=StrictUndefined,
    )
    for f in j2_files:
        try:
            env.get_template(f)
        except Exception as e:
            problems.append(f'模板 {f} Jinja2 编译失败: {e}')


def _check_excel(tpl_dir, problems):
    """4. Excel 数据健康"""
    import openpyxl
    excel_dir = os.path.join(tpl_dir, 'excel')
    if not os.path.isdir(excel_dir):
        return
    for f in sorted(os.listdir(excel_dir)):
        if not f.lower().endswith(('.xlsx', '.xls')):
            continue
        try:
            openpyxl.load_workbook(os.path.join(excel_dir, f), read_only=True, data_only=True)
        except Exception as e:
            problems.append(f'Excel {f} 无法读取: {e}')


# ---- 5.0.1（501-b）：参数合理性 + 协议兼容性检查 ----

# 端口速率解析（'400G'→400，'1G'→1，'10G'→10 …）
_RATE_RE = re.compile(r'^(\d+(?:\.\d+)?)\s*[Gg]$')


def _rate_gbps(rate):
    m = _RATE_RE.match(str(rate or '').strip())
    return float(m.group(1)) if m else None


def _read_param_table(tpl_dir):
    """读取 excel/parameter.xlsx 参数表 → {全局参数名: 参数值}（仅 key=列名 参数表）。"""
    import openpyxl
    path = os.path.join(tpl_dir, 'excel', 'parameter.xlsx')
    if not os.path.exists(path):
        return {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    out = {}
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            continue
        # 参数表形如：全局参数名/全局参数名称 | 参数值
        kcol = vcol = None
        for i, h in enumerate(header):
            s = str(h or '').strip()
            if s in ('全局参数名', '全局参数名称'):
                kcol = i
            elif s == '参数值':
                vcol = i
        if kcol is None or vcol is None:
            continue
        for row in rows:
            k, v = row[kcol], row[vcol]
            if k is None:
                continue
            out[str(k).strip()] = v
    wb.close()
    return out


def _resolve_device(device_id):
    """设备 id → 设备 dict（backend 设备库；导入失败返回 None）。"""
    try:
        sys.path.insert(0, BACKEND)
        from intent.device_library import get_device
        return get_device(device_id)
    except Exception:  # noqa: BLE001
        return None


def _load_plan(tpl_dir):
    """读取 plan.json；无/损坏返回 None。"""
    path = os.path.join(tpl_dir, 'plan.json')
    if not os.path.exists(path):
        return None
    try:
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    except (OSError, ValueError):
        return None


def _check_param_reasonableness(tpl_dir, problems):
    """6. 参数合理性：PFC/CNP 队列 0-7、收敛比、BGP AS 段、VLAN/IP 段重叠、端口速率与设备库匹配、参数表键完整。"""
    import ipaddress

    qv = _read_param_table(tpl_dir)
    # 6a. 参数表键完整性：template.meta.json 声明的 tunables 须出现在参数表
    meta_path = os.path.join(tpl_dir, 'template.meta.json')
    tunables = []
    if os.path.exists(meta_path):
        try:
            with open(meta_path, encoding='utf-8') as f:
                tunables = json.load(f).get('tunables') or []
        except (OSError, ValueError):
            pass
    missing_keys = [k for k in tunables if k not in qv]
    if missing_keys:
        problems.append(f'参数表缺可调参数键: {missing_keys}')

    # 6b. PFC/CNP 队列 0-7（参数表源）
    for k in ('PFC队列', 'CNP队列'):
        if k in qv:
            try:
                v = int(qv[k])
            except (TypeError, ValueError):
                problems.append(f'参数表 {k} 非整数: {qv[k]!r}')
                continue
            if not (0 <= v <= 7):
                problems.append(f'参数表 {k} 须在 0-7: {v}')

    # 以下 plan 级检查仅对带 plan.json 的模板（AIDC 四表格）执行；老式模板跳过
    plan = _load_plan(tpl_dir)
    if plan is None:
        return
    macro = plan.get('macro', {})

    # 6c. 收敛比：IB=1 / RoCE 合理区间 [1,8]
    from intent.device_library import resolve_models_fabric
    fabric = resolve_models_fabric(macro.get('deviceModels') or {})
    conv = macro.get('convergence')
    if conv is not None:
        try:
            cv = float(conv)
        except (TypeError, ValueError):
            problems.append(f'convergence 非数值: {conv!r}')
        else:
            if fabric == 'ib' and cv != 1:
                problems.append(f'IB 收敛比须为 1（1:1），实际 {conv}')
            elif fabric == 'roce' and not (1 <= cv <= 8):
                problems.append(f'RoCE 收敛比须在合理区间 [1,8]，实际 {conv}')

    # 6d. BGP AS 段冲突：设备 ASN 须落在 asRange 内
    as_range = macro.get('asRange') or [65001, 65500]
    if len(as_range) == 2 and as_range[0] <= as_range[1]:
        for d in plan.get('deviceList', []):
            asn = d.get('asn')
            if asn is not None:
                try:
                    if not (int(as_range[0]) <= int(asn) <= int(as_range[1])):
                        problems.append(f'设备 {d.get("name")} AS {asn} 越出 asRange {as_range}')
                except (TypeError, ValueError):
                    problems.append(f'设备 {d.get("name")} AS 非法: {asn!r}')

    # 6e. VLAN 段重叠
    vlan_ranges = macro.get('vlanRanges') or {}
    planes = [(k, v) for k, v in vlan_ranges.items() if isinstance(v, (list, tuple)) and len(v) == 2]
    for i in range(len(planes)):
        for j in range(i + 1, len(planes)):
            (n1, (a1, b1)), (n2, (a2, b2)) = planes[i], planes[j]
            try:
                overlap = not (int(b1) < int(a2) or int(b2) < int(a1))
            except (TypeError, ValueError):
                continue
            if overlap:
                problems.append(f'VLAN 段重叠: {n1}[{a1},{b1}] 与 {n2}[{a2},{b2}]')

    # 6f. IP 段重叠（macro.ipSegments）
    ip_segments = macro.get('ipSegments') or {}
    segs = []
    for k, v in ip_segments.items():
        try:
            segs.append((k, ipaddress.ip_network(str(v))))
        except ValueError:
            problems.append(f'ipSegments.{k} 非法网段: {v!r}')
    for i in range(len(segs)):
        for j in range(i + 1, len(segs)):
            (n1, net1), (n2, net2) = segs[i], segs[j]
            if net1.overlaps(net2):
                problems.append(f'IP 段重叠: {n1}({net1}) 与 {n2}({net2})')

    # 6g. 端口速率与设备库匹配：plan 声明型号 → 设备库端口速率（多速率取 port_speed_max 上限）；
    #     连接速率须 ≤ 型号最大端口速率
    models = macro.get('deviceModels') or {}
    for c in plan.get('connections', []):
        rate = _rate_gbps(c.get('rate'))
        if rate is None:
            continue
        dev = next((d for d in plan.get('deviceList', []) if d.get('name') == c.get('src')), None)
        role = (dev or {}).get('role', '')
        m = models.get(role)
        if not m:
            continue
        did = _lookup_device_id(m)
        dev_spec = _resolve_device(did) if did else None
        if not dev_spec:
            continue
        cap = max(_rate_gbps(dev_spec.get('port_speed')) or 0,
                  _rate_gbps(dev_spec.get('port_speed_max')) or 0)
        if cap and rate > cap:
            problems.append(f'连接速率越限: {c.get("src")} {c.get("src_port")} {c.get("rate")}'
                            f' > 型号 {m} 最大端口速率 {dev_spec.get("port_speed_max", dev_spec.get("port_speed"))}')


def _lookup_device_id(model):
    try:
        sys.path.insert(0, BACKEND)
        from intent.device_library import lookup_device_by_model
        return lookup_device_by_model(model)
    except Exception:  # noqa: BLE001
        return None


def _check_protocol_compat(tpl_dir, problems):
    """7. 协议兼容性：plan/参数表设备型号与协议匹配（IB 用 IB 交换机 / RoCE 用 RoCE 交换机）、模板与 plan 一致。"""
    plan = _load_plan(tpl_dir)
    if plan is None:
        return
    macro = plan.get('macro', {})
    models = macro.get('deviceModels') or {}
    fabric = _resolve_fabric(models)
    _PARAM_STO = ('SPINE', 'LEAF', 'STO_SPINE', 'STO_LEAF')

    # 7a. plan 声明型号 → 设备库解析 + 参数/存储角色协议与 plan fabric 一致
    for role, m in models.items():
        if not m:
            continue
        did = _lookup_device_id(m)
        if did is None:
            problems.append(f'plan 型号 {role}={m} 不在设备库')
            continue
        dev = _resolve_device(did)
        proto = (dev or {}).get('protocol')
        if role in _PARAM_STO and proto and proto != fabric:
            problems.append(f'协议不匹配: {role}={m}({proto}) 与 plan fabric({fabric}) 不符')

    # 7b. 模板与 plan.json 协议一致：hostname.xlsx 参数/存储平面 型号协议须与 plan fabric 一致
    import openpyxl
    host_path = os.path.join(tpl_dir, 'excel', 'hostname.xlsx')
    if os.path.exists(host_path):
        seen_protos = set()
        try:
            wb = openpyxl.load_workbook(host_path, read_only=True, data_only=True)
        except Exception:
            wb = None
        if wb is not None:
            for ws in wb.worksheets:
                if not (ws.title.startswith('设备表-参数网') or ws.title.startswith('设备表-存储网')):
                    continue
                rows = ws.iter_rows(values_only=True)
                next(rows, None)  # header
                for row in rows:
                    model = row[1] if len(row) > 1 else None
                    if not model:
                        continue
                    did = _lookup_device_id(model)
                    dev = _resolve_device(did) if did else None
                    proto = (dev or {}).get('protocol')
                    if proto:
                        seen_protos.add(proto)
            wb.close()
        if seen_protos and seen_protos != {fabric}:
            problems.append(f'hostname.xlsx 参数/存储型号协议 {sorted(seen_protos)} 与 plan fabric({fabric}) 不一致')


def _resolve_fabric(models):
    try:
        sys.path.insert(0, BACKEND)
        from intent.device_library import resolve_models_fabric
        return resolve_models_fabric(models)
    except Exception:  # noqa: BLE001
        return 'roce'


def _render_all(templates, tmpdir):
    """在临时 workspace 中一次性 dry-run 渲染全部模板，返回 {template: [results]}。"""
    import pandas as pd
    import io
    import contextlib

    os.environ['MC_WORKSPACE'] = tmpdir
    sys.path.insert(0, BACKEND)
    logging.getLogger('magiccommander').setLevel(logging.ERROR)

    for name in templates:
        shutil.copytree(os.path.join(base, name), os.path.join(tmpdir, name))
    pd.DataFrame({'项目名称': templates}).to_excel(
        os.path.join(tmpdir, 'MC_Para.xlsx'), sheet_name='项目名称', index=False, header=True)

    from pre_processing import PreProcessing
    logging.getLogger().setLevel(logging.ERROR)  # 屏蔽后端 INFO（pre_processing 日志器）

    p = PreProcessing()
    p.read_MC_para('MC_Para.xlsx')
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        p.execute_dry_run('all', 'device_name')

    last = None
    for line in buf.getvalue().splitlines():
        line = line.strip()
        if line.startswith('{'):
            try:
                last = json.loads(line)
            except (ValueError, TypeError):
                continue
    if not last or last.get('status') != 'complete':
        raise RuntimeError(f'渲染未完成: {str(last)[:300]}')

    grouped: dict = {}
    for r in (last.get('data') or {}).get('results') or []:
        grouped.setdefault(r.get('project', ''), []).append(r)
    for name in templates:
        grouped.setdefault(name, [])
    return grouped


def main():
    templates = _discover_templates()
    print(f'共发现 {len(templates)} 个模板\n')

    render_results: dict = {}
    if templates:
        tmpdir = tempfile.mkdtemp(prefix='mc_tpl_')
        try:
            render_results = _render_all(templates, tmpdir)
        except Exception as e:
            render_results = {'__error__': f'{type(e).__name__}: {e}'}
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)

    failures = 0
    for t in templates:
        tpl_dir = os.path.join(base, t)
        problems = []
        j2_files = _check_structure(tpl_dir, problems)
        _check_meta(tpl_dir, problems)
        if j2_files:
            _check_jinja(tpl_dir, j2_files, problems)
        _check_excel(tpl_dir, problems)

        # 6/7. 参数合理性 + 协议兼容性（5.0.1 / 501-b）
        _check_param_reasonableness(tpl_dir, problems)
        _check_protocol_compat(tpl_dir, problems)

        # 8. 渲染健康
        if '__error__' in render_results:
            problems.append(f'渲染引擎异常: {render_results["__error__"]}')
        else:
            results = render_results.get(t, [])
            if not results:
                problems.append('渲染产出 0 台设备（para.xlsx 与 excel 数据可能为空）')
            errors = [r for r in results if isinstance(r.get('content'), str) and r['content'].startswith('[渲染错误]')]
            if errors:
                problems.append(f'存在渲染错误设备 {len(errors)} 台，如: {errors[0].get("device")}')

        ok = not problems
        if not ok:
            failures += 1
        print(f'[{"OK" if ok else "FAIL"}] {t}: 模板 {len(j2_files) if j2_files else 0} 个, 结构/元数据/语法/数据/参数合理性/协议兼容性/渲染健康')
        for p in problems:
            print(f'       - {p}')

    print(f'\n结果: {len(templates) - failures}/{len(templates)} 模板通过')
    sys.exit(1 if failures else 0)


if __name__ == '__main__':
    main()
