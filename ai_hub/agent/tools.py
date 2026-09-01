"""
Agent Tool 定义
将现有 Python CLI 功能包装为标准 Tool 接口，供 LLM 调用
"""
import json
import logging
import os
import subprocess
import sys
from pathlib import Path
from typing import Any

from ai_hub.agent.schemas import ToolPermission, get_tool_permission
from ai_hub.agent.preset_templates import TEMPLATE_MAP, _TPL_GENERIC

logger = logging.getLogger(__name__)

# 工具注册表
_tools: dict[str, dict] = {}


def register_tool(name: str, description: str, parameters: dict, handler: callable,
                  permission: ToolPermission | None = None):
    """注册一个 Agent Tool"""
    _tools[name] = {
        "name": name,
        "description": description,
        "parameters": parameters,
        "handler": handler,
        "permission": permission.value if permission else get_tool_permission(name).value,
    }


def get_tool_definitions() -> list[dict]:
    """获取所有工具定义（JSON Schema 格式），供 LLM function calling 使用"""
    return [
        {
            "type": "function",
            "function": {
                "name": t["name"],
                "description": t["description"],
                "parameters": t["parameters"],
                "permission": t.get("permission", "confirm"),
            },
        }
        for t in _tools.values()
    ]


async def execute_tool(name: str, arguments: dict) -> dict:
    """执行指定工具（4.3 F3-4：参数校验 + 业务错误可读化，全部失败均返回可读中文错误）"""
    tool = _tools.get(name)
    if not tool:
        return {"success": False, "error": f"未知工具: {name}"}
    # 参数校验（必需字段缺失/类型错误/enum 越界 → 可读中文错误，不抛异常）
    errors = _validate_tool_args(name, arguments, tool.get("parameters", {}))
    if errors:
        return {"success": False, "error": "；".join(errors)}
    try:
        result = await tool["handler"](arguments)
        # 业务错误可读化：handler 返回 {"status":"error", "error":...} JSON 时转为失败
        if isinstance(result, str):
            try:
                parsed = json.loads(result)
                if isinstance(parsed, dict) and parsed.get("status") == "error":
                    err = parsed.get("error") or parsed.get("message") or "操作失败"
                    return {"success": False, "error": err}
            except Exception:
                pass
        return {"success": True, "result": result}
    except Exception as e:
        logger.error(f"Tool '{name}' execution failed: {e}")
        return {"success": False, "error": str(e)}


def _validate_tool_args(name: str, args: dict, schema: dict) -> list[str]:
    """按注册的工具 JSON Schema 校验参数：必需字段缺失、类型错误、enum 越界。

    返回可读中文错误列表（空列表表示通过）。
    """
    errors: list[str] = []
    props = schema.get("properties", {}) or {}
    required = schema.get("required", []) or []
    for r in required:
        if r not in args or args[r] is None or args[r] == "":
            errors.append(f"工具 {name} 缺少必需参数: {r}")
    for key, val in args.items():
        prop = props.get(key)
        if not prop:
            continue
        ptype = prop.get("type")
        if ptype == "string" and not isinstance(val, str):
            errors.append(f"参数 {key} 应为字符串")
        elif ptype == "integer" and not isinstance(val, int):
            errors.append(f"参数 {key} 应为整数")
        elif ptype == "number" and not isinstance(val, (int, float)):
            errors.append(f"参数 {key} 应为数字")
        if prop.get("enum") and val not in prop["enum"]:
            errors.append(f"参数 {key} 取值无效，可选: {', '.join(str(e) for e in prop['enum'])}")
    return errors


async def _run_python_cli(args: list[str]) -> str:
    """异步运行 Python CLI 命令并返回输出（不阻塞 uvicorn 事件循环，显式 UTF-8 解码防中文乱码）"""
    import asyncio
    workspace = _workspace_dir or ""
    backend_dir = _backend_dir or str(Path(__file__).parent.parent.parent / "backend")

    env = {**__import__("os").environ}
    if workspace:
        env["MC_WORKSPACE"] = workspace

    try:
        proc = await asyncio.create_subprocess_exec(
            sys.executable, str(Path(backend_dir) / "main.py"), *args,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=backend_dir,
            env=env,
        )
        try:
            stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=120)
        except asyncio.TimeoutError:
            proc.kill()
            await proc.wait()
            return "命令执行超时"
        output = stdout.decode("utf-8", errors="replace").strip()
        if not output:
            output = stderr.decode("utf-8", errors="replace").strip()
        # 过滤进度消息，只提取最终状态行（避免进度JSON污染LLM上下文）
        lines = output.split('\n')
        for line in reversed(lines):
            line = line.strip()
            if not line:
                continue
            try:
                import json as _json
                data = _json.loads(line)
                if isinstance(data, dict) and data.get('status') in ('success', 'complete', 'error'):
                    # 只返回最终状态行（进度行 status='progress' 被忽略）
                    return line
            except Exception:
                pass
        return output  # 兜底：无状态行时返回原始输出
    except Exception as e:
        return f"命令执行失败: {e}"


_workspace_dir = ""
_backend_dir = ""


def set_workspace_dir(path: str):
    global _workspace_dir
    _workspace_dir = path


def set_backend_dir(path: str):
    global _backend_dir
    _backend_dir = path


# ====== 注册所有工具 ======


async def _list_projects(args: dict) -> str:
    return await _run_python_cli(["project", "list"])


async def _create_project(args: dict) -> str:
    project_name = args["projectName"]
    cmd = ["project", "create", project_name]
    if args.get("templateName"):
        cmd.extend(["--template", args["templateName"]])
    return await _run_python_cli(cmd)


async def _render_config(args: dict) -> str:
    project_name = args["projectName"]
    return await _run_python_cli(["render", "project", project_name])


async def _dry_run(args: dict) -> str:
    project_name = args["projectName"]
    return await _run_python_cli(["render", "dry-run", project_name])


async def _validate_template(args: dict) -> str:
    template_name = args["templateName"]
    return await _run_python_cli(["validate", "template", template_name])


async def _validate_excel(args: dict) -> str:
    project_name = args["projectName"]
    return await _run_python_cli(["validate", "excel", project_name])


async def _diff_compare(args: dict) -> str:
    project_name = args["projectName"]
    return await _run_python_cli(["diff", "compare", project_name])


async def _read_file(args: dict) -> str:
    project_name = args["projectName"]
    file_path = args["filePath"]
    return await _run_python_cli(["project", "read-file", project_name, file_path])


async def _search_files(args: dict) -> str:
    """搜索项目文件：通过 list-files + grep 实现"""
    query = args["query"]
    project_name = args.get("projectName", "")
    ws = _workspace_dir or "workspace"
    import glob as _glob

    results = []
    if project_name:
        search_dirs = [str(Path(ws) / project_name)]
    else:
        search_dirs = [str(Path(ws) / d) for d in os.listdir(ws)
                       if os.path.isdir(os.path.join(ws, d)) and not d.startswith('.')]

    for search_dir in search_dirs:
        for root, dirs, files in os.walk(search_dir):
            dirs[:] = [d for d in dirs if not d.startswith('.') and d != '__pycache__']
            for fname in files:
                if query.lower() in fname.lower():
                    results.append(os.path.relpath(os.path.join(root, fname), ws))
                else:
                    fpath = os.path.join(root, fname)
                    try:
                        with open(fpath, 'r', encoding='utf-8', errors='ignore') as f:
                            for i, line in enumerate(f, 1):
                                if query.lower() in line.lower():
                                    results.append(f"{os.path.relpath(fpath, ws)}:{i}: {line.strip()[:200]}")
                                    break
                    except Exception:
                        pass

    if not results:
        return json.dumps({"status": "ok", "message": f"未找到匹配 '{query}' 的文件或内容", "data": []}, ensure_ascii=False)
    return json.dumps({"status": "ok", "message": f"找到 {len(results)} 个匹配项", "data": results[:50]}, ensure_ascii=False)


async def _create_template(args: dict) -> str:
    source = args["sourceProject"]
    name = args["templateName"]
    return await _run_python_cli(["template", "save", source, name])


async def _list_templates(args: dict) -> str:
    return await _run_python_cli(["template", "list"])


async def _delete_template(args: dict) -> str:
    name = args["templateName"]
    return await _run_python_cli(["template", "delete", "--force", name])


async def _update_template(args: dict) -> str:
    name = args["templateName"]
    file_path = args["filePath"]
    content = args["content"]
    return await _run_python_cli(["template", "update", name, file_path, content])


async def _create_project_intelligent(args: dict) -> str:
    """智能创建项目：根据设备类型和需求自动生成模板和参数表"""
    project_name = args["projectName"]
    device_type = args.get("deviceType", "switch")
    vendor = args.get("vendor", "huawei")
    config_description = args.get("configDescription", "")

    ws = _workspace_dir or "workspace"
    project_dir = Path(ws) / project_name

    if project_dir.exists():
        return json.dumps({
            "error": f"项目 '{project_name}' 已存在",
            "status": "exists",
        }, ensure_ascii=False)

    # 创建目录结构
    (project_dir / "templates").mkdir(parents=True, exist_ok=True)
    (project_dir / "excel").mkdir(parents=True, exist_ok=True)
    (project_dir / "output").mkdir(parents=True, exist_ok=True)
    (project_dir / "output-label").mkdir(parents=True, exist_ok=True)
    (project_dir / "yaml").mkdir(parents=True, exist_ok=True)

    # 生成 Jinja2 模板
    template_content = _generate_template(device_type, vendor, config_description)
    template_name = f"{device_type.upper()}.j2"
    template_path = project_dir / "templates" / template_name
    template_path.write_text(template_content, encoding="utf-8")

    # 生成 Excel 参数文件（使用 openpyxl 直接创建）
    _create_excel_from_template(project_dir, device_type, vendor)

    result = {
        "status": "created",
        "projectName": project_name,
        "structure": {
            "templates": [template_name],
            "directories": ["templates", "excel", "output", "output-label", "yaml"],
        },
        "templatePreview": template_content[:500],
        "message": f"项目 '{project_name}' 已创建，包含 {device_type} 类型的 {vendor} 配置模板。",
    }
    return json.dumps(result, ensure_ascii=False)


def _generate_template(device_type: str, vendor: str, description: str) -> str:
    """根据设备类型和厂商生成 Jinja2 模板"""
    vendor_upper = vendor.upper()

    tpl = TEMPLATE_MAP.get((device_type, vendor.lower()))
    if not tpl:
        tpl = _TPL_GENERIC.format(
            device_type=device_type,
            vendor=vendor_upper,
            description=description or f"{vendor_upper} {device_type} 配置模板",
        )

    return tpl


def _create_excel_from_template(project_dir: Path, device_type: str, vendor: str):
    """使用 openpyxl 创建基础 Excel 参数文件"""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not available, skipping Excel creation")
        return

    excel_path = project_dir / "excel" / "parameter.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "主机表"

    # 根据设备类型定义列
    columns = {
        "switch": ["设备名", "管理IP", "掩码", "管理接口", "VLAN", "网关IP", "网关掩码", "网关接口",
                    "SNMP团体名", "SNMP地址", "NTP地址", "LOGHOST地址",
                    "AAA名称", "AAA地址", "NAS_IP", "AAA认证密钥",
                    "domain名称", "本地用户名", "本地用户密钥"],
        "router": ["设备名", "管理IP", "掩码", "管理接口", "路由协议", "AS号",
                    "SNMP团体名", "SNMP地址", "NTP地址", "LOGHOST地址",
                    "AAA名称", "AAA地址", "NAS_IP", "AAA认证密钥",
                    "本地用户名", "本地用户密钥"],
        "firewall": ["设备名", "管理IP", "掩码", "管理接口", "安全域", "策略名称",
                      "SNMP团体名", "SNMP地址", "NTP地址", "LOGHOST地址",
                      "AAA名称", "AAA地址", "本地用户名", "本地用户密钥"],
    }

    headers = columns.get(device_type, columns["switch"])
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 添加示例行
    sample_row = [f"示例设备-{i}" for i in range(1, len(headers) + 1)]
    for col, val in enumerate(sample_row, 1):
        ws.cell(row=2, column=col, value=val)

    wb.save(str(excel_path))
    logger.info(f"Created Excel parameter file: {excel_path}")


async def _reverse_engineer_config(args: dict) -> str:
    """从已有网络设备配置反推模板和参数表"""
    config_text = args["configText"]
    project_name = args["projectName"]
    vendor = args.get("vendor", "huawei")
    device_type = args.get("deviceType", "switch")

    ws = _workspace_dir or "workspace"
    project_dir = Path(ws) / project_name

    if project_dir.exists():
        return json.dumps({
            "error": f"项目 '{project_name}' 已存在，请使用其他名称",
            "status": "exists",
        }, ensure_ascii=False)

    # 提取变量
    import re
    extracted = {}

    # 提取设备名
    hostname_patterns = [
        (r'^sysname\s+(\S+)', '设备名'),
        (r'^hostname\s+(\S+)', '设备名'),
    ]
    for pattern, key in hostname_patterns:
        m = re.search(pattern, config_text, re.MULTILINE)
        if m:
            extracted[key] = m.group(1)
            break

    # 提取 IP 地址（管理接口）
    ip_pattern = r'interface\s+\S+\s*\n\s*ip\s+address\s+(\d+\.\d+\.\d+\.\d+)\s+(\d+\.\d+\.\d+\.\d+)'
    m = re.search(ip_pattern, config_text)
    if m:
        extracted["管理IP"] = m.group(1)
        extracted["掩码"] = m.group(2)

    # 提取 VLAN
    vlan_pattern = r'vlan\s+(\d+)'
    m = re.search(vlan_pattern, config_text, re.IGNORECASE)
    if m:
        extracted["VLAN"] = m.group(1)

    # 提取网关 IP
    gw_pattern = r'Vlanif\d+\s*\n.*\n\s*ip\s+address\s+(\d+\.\d+\.\d+\.\d+)\s+(\d+\.\d+\.\d+\.\d+)'
    m = re.search(gw_pattern, config_text)
    if not m:
        gw_pattern = r'interface\s+Vlan\S*\s*\n.*\n\s*ip\s+address\s+(\d+\.\d+\.\d+\.\d+)\s+(\d+\.\d+\.\d+\.\d+)'
        m = re.search(gw_pattern, config_text)
    if m:
        extracted["网关IP"] = m.group(1)
        extracted["网关掩码"] = m.group(2)

    # 提取 SNMP
    snmp_patterns = [
        r'snmp-agent\s+community\s+read\s+(\S+)',
        r'snmp-server\s+community\s+(\S+)\s+RO',
        r'snmp-server\s+community\s+(\S+)',
    ]
    for p in snmp_patterns:
        m = re.search(p, config_text)
        if m:
            extracted["SNMP团体名"] = m.group(1)
            break

    snmp_host_patterns = [
        r'snmp-agent\s+target-host\s+trap\s+address\s+udp-domain\s+(\S+)',
        r'snmp-server\s+host\s+(\S+)',
    ]
    for p in snmp_host_patterns:
        m = re.search(p, config_text)
        if m:
            extracted["SNMP地址"] = m.group(1)
            break

    # 提取 NTP
    ntp_patterns = [
        r'ntp-service\s+unicast-server\s+(\S+)',
        r'ntp\s+server\s+(\S+)',
    ]
    for p in ntp_patterns:
        m = re.search(p, config_text)
        if m:
            extracted["NTP地址"] = m.group(1)
            break

    # 提取日志服务器
    log_patterns = [
        r'info-center\s+loghost\s+(\S+)',
        r'logging\s+host\s+(\S+)',
    ]
    for p in log_patterns:
        m = re.search(p, config_text)
        if m:
            extracted["LOGHOST地址"] = m.group(1)
            break

    # 提取 AAA
    tacacs_patterns = [
        r'hwtacacs\s+scheme\s+(\S+)',
        r'tacacs-server\s+host\s+(\S+)',
    ]
    for p in tacacs_patterns:
        m = re.search(p, config_text)
        if m:
            if "AAA名称" not in extracted:
                extracted["AAA名称"] = m.group(1)
            else:
                extracted.setdefault("AAA地址", m.group(1))

    aaa_key_pattern = r'key\s+\S+\s+simple\s+(\S+)'
    m = re.search(aaa_key_pattern, config_text)
    if m:
        extracted["AAA认证密钥"] = m.group(1)

    aaa_ip_pattern = r'primary\s+\S+\s+(\S+)'
    m = re.search(aaa_ip_pattern, config_text)
    if m:
        extracted["AAA地址"] = m.group(1)

    nas_pattern = r'nas-ip\s+(\S+)'
    m = re.search(nas_pattern, config_text)
    if m:
        extracted["NAS_IP"] = m.group(1)

    # 提取 domain
    domain_pattern = r'domain\s+(\S+)\s*\n'
    m = re.search(domain_pattern, config_text)
    if m:
        extracted["domain名称"] = m.group(1)

    # 提取本地用户
    user_patterns = [
        r'local-user\s+(\S+)\s+class\s+manage',
        r'username\s+(\S+)\s+privilege',
    ]
    for p in user_patterns:
        m = re.search(p, config_text)
        if m:
            extracted["本地用户名"] = m.group(1)
            break

    pass_patterns = [
        r'password\s+simple\s+(\S+)',
        r'password\s+(\S+)',
        r'secret\s+(\S+)',
    ]
    for p in pass_patterns:
        m = re.search(p, config_text)
        if m:
            extracted["本地用户密钥"] = m.group(1)
            break

    # 提取管理接口
    mgmt_pattern = r'interface\s+(\S+)\s*\n\s*ip\s+address\s+'
    m = re.search(mgmt_pattern, config_text)
    if m:
        extracted["管理接口"] = m.group(1)

    # 提取 VLAN 网关接口
    gw_if_pattern = r'interface\s+(Vlanif\d+|Vlan\d+)\s*\n'
    m = re.search(gw_if_pattern, config_text)
    if m:
        extracted["网关接口"] = m.group(1)

    if not extracted:
        return json.dumps({
            "status": "error",
            "error": "未能从配置文本中提取到有效参数。请确认配置文本格式正确。",
        }, ensure_ascii=False)

    # 创建项目目录
    (project_dir / "templates").mkdir(parents=True, exist_ok=True)
    (project_dir / "excel").mkdir(parents=True, exist_ok=True)
    (project_dir / "output").mkdir(parents=True, exist_ok=True)
    (project_dir / "output-label").mkdir(parents=True, exist_ok=True)
    (project_dir / "yaml").mkdir(parents=True, exist_ok=True)

    # 生成模板（替换提取的值为 Jinja2 变量）
    template_content = config_text
    variable_map = []

    # 按长度降序排序，避免短字符串先替换导致长字符串被破坏
    replacements = []
    for key, value in sorted(extracted.items(), key=lambda x: -len(x[1])):
        var_name = key
        if value and value in template_content:
            template_content = template_content.replace(value, f"{{{{ info['{var_name}'] }}}}")
            replacements.append({"变量名": var_name, "原值": value})

    # 写模板文件
    template_name = f"{device_type.upper()}_reversed.j2"
    template_path = project_dir / "templates" / template_name
    template_path.write_text(
        f"{{# 从配置反向生成 - {project_name} #}}\n{template_content}",
        encoding="utf-8",
    )

    # 创建 Excel 参数表
    try:
        import openpyxl
        excel_path = project_dir / "excel" / "parameter.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "主机表"

        keys = list(extracted.keys())
        for col, key in enumerate(keys, 1):
            ws.cell(row=1, column=col, value=key)
        for col, key in enumerate(keys, 1):
            ws.cell(row=2, column=col, value=extracted.get(key, ""))

        # 反向生成替换明细表
        ws2 = wb.create_sheet("替换明细")
        ws2.cell(row=1, column=1, value="变量名")
        ws2.cell(row=1, column=2, value="原值")
        for row, item in enumerate(replacements, 2):
            ws2.cell(row=row, column=1, value=item["变量名"])
            ws2.cell(row=row, column=2, value=item["原值"])

        wb.save(str(excel_path))
    except ImportError:
        logger.warning("openpyxl not available, skipping Excel creation")

    result = {
        "status": "created",
        "projectName": project_name,
        "extractedVariables": {k: v for k, v in list(extracted.items())[:10]},
        "totalVariables": len(extracted),
        "templateName": template_name,
        "message": f"从配置文本反向生成了项目 '{project_name}'，提取了 {len(extracted)} 个变量。模板已保存为 {template_name}。",
    }
    return json.dumps(result, ensure_ascii=False)


async def _recommend_template(args: dict) -> str:
    """根据项目特征推荐合适的模板"""
    device_type = args.get("deviceType", "")
    vendor = args.get("vendor", "")
    project_name = args.get("projectName", "")

    # 预置模板目录
    template_catalog = {
        "华为交换机": {
            "deviceType": "switch", "vendor": "huawei",
            "template": "HUAWEI_SWITCH", "description": "华为交换机基础配置模板，包含管理接口、VLAN、SNMP、NTP、AAA、本地用户",
            "features": ["VLAN配置", "聚合接口", "SNMP", "NTP", "AAA/TACACS", "SSH"],
        },
        "思科交换机": {
            "deviceType": "switch", "vendor": "cisco",
            "template": "CISCO_SWITCH", "description": "思科交换机基础配置模板，包含管理接口、VLAN、SNMP、NTP、AAA、本地用户",
            "features": ["VLAN配置", "SNMP", "NTP", "TACACS+", "SSH"],
        },
        "H3C交换机": {
            "deviceType": "switch", "vendor": "h3c",
            "template": "H3C_SWITCH", "description": "H3C交换机基础配置模板，包含管理接口、VLAN、SNMP、NTP、AAA、本地用户",
            "features": ["VLAN配置", "SNMP", "NTP", "AAA/TACACS", "SSH"],
        },
        "华为路由器": {
            "deviceType": "router", "vendor": "huawei",
            "template": "HUAWEI_ROUTER", "description": "华为路由器基础配置模板，支持OSPF/BGP路由协议",
            "features": ["OSPF", "BGP", "SNMP", "NTP", "AAA", "SSH"],
        },
        "思科路由器": {
            "deviceType": "router", "vendor": "cisco",
            "template": "CISCO_ROUTER", "description": "思科路由器基础配置模板，支持OSPF/BGP路由协议",
            "features": ["OSPF", "BGP", "SNMP", "NTP", "AAA", "SSH"],
        },
        "华为防火墙": {
            "deviceType": "firewall", "vendor": "huawei",
            "template": "HUAWEI_FIREWALL", "description": "华为防火墙基础配置模板，包含安全域、安全策略",
            "features": ["安全域", "安全策略", "SNMP", "AAA", "SSH"],
        },
    }

    # 根据用户输入匹配
    recommendations = []
    for name, info in template_catalog.items():
        score = 0
        if device_type and info["deviceType"] == device_type:
            score += 3
        if vendor and info["vendor"] == vendor:
            score += 3
        if not device_type and not vendor:
            score += 1  # 如果没有指定条件，显示所有模板

        if score > 0:
            recommendations.append({
                "name": name,
                "score": score,
                "deviceType": info["deviceType"],
                "vendor": info["vendor"],
                "description": info["description"],
                "features": info["features"],
            })

    # 按匹配度排序
    recommendations.sort(key=lambda x: -x["score"])

    # 如果指定了项目名，分析项目现有模板
    project_analysis = None
    if project_name:
        ws = _workspace_dir or "workspace"
        project_dir = Path(ws) / project_name
        if project_dir.exists():
            templates_dir = project_dir / "templates"
            if templates_dir.exists():
                existing = list(templates_dir.glob("*.j2"))
                if existing:
                    project_analysis = {
                        "existingTemplates": [f.name for f in existing],
                        "suggestion": "现有模板可基于推荐模板进行对比和优化",
                    }

    result = {
        "status": "ok",
        "recommendations": recommendations[:5],
        "totalAvailable": len(template_catalog),
        "projectAnalysis": project_analysis,
        "message": f"找到 {len(recommendations)} 个匹配的模板推荐" if recommendations else "未找到匹配的模板，请尝试指定设备类型和厂商",
    }
    return json.dumps(result, ensure_ascii=False)


# ====== 新增工具：补齐 CLI 能力 ======

async def _delete_project(args: dict) -> str:
    project_name = args["projectName"]
    return await _run_python_cli(["project", "delete", "--force", project_name])


async def _get_project_info(args: dict) -> str:
    project_name = args["projectName"]
    return await _run_python_cli(["project", "info", "--format", "json", project_name])


async def _render_yaml(args: dict) -> str:
    project_name = args["projectName"]
    return await _run_python_cli(["render", "yaml", project_name])


async def _undo_render(args: dict) -> str:
    project_name = args["projectName"]
    return await _run_python_cli(["render", "undo", project_name])


async def _generate_labels(args: dict) -> str:
    project_name = args["projectName"]
    return await _run_python_cli(["label", "print", project_name])


async def _generate_label_md(args: dict) -> str:
    project_name = args["projectName"]
    return await _run_python_cli(["label", "md", project_name])


async def _delete_labels(args: dict) -> str:
    project_name = args["projectName"]
    return await _run_python_cli(["label", "delete", project_name])


async def _delete_files(args: dict) -> str:
    """删除项目输出文件（清空渲染结果）"""
    project_name = args["projectName"]
    file_type = args.get("fileType", "output")
    return await _run_python_cli(["file", "delete", "--force", file_type, project_name])


async def _list_project_files(args: dict) -> str:
    project_name = args["projectName"]
    return await _run_python_cli(["project", "list-files", project_name])


async def _read_excel(args: dict) -> str:
    project_name = args["projectName"]
    file_name = args["fileName"]
    cmd = ["project", "read-excel", project_name, file_name]
    if args.get("sheetName"):
        cmd.extend(["--sheet", args["sheetName"]])
    return await _run_python_cli(cmd)


async def _write_excel(args: dict) -> str:
    project_name = args["projectName"]
    file_name = args["fileName"]
    data = json.dumps(args["data"], ensure_ascii=False)
    return await _run_python_cli(["project", "write-excel", project_name, file_name, data])


async def _write_text_file(args: dict) -> str:
    project_name = args["projectName"]
    file_path = args["filePath"]
    content = args["content"]
    return await _run_python_cli(["project", "write-file", project_name, file_path, content])


async def _analyze_project(args: dict) -> str:
    project_name = args["projectName"]
    return await _run_python_cli(["analyze", "project", project_name])


# ====== 4.3 F3-4：项目/模板操作工具补齐 ======

async def _update_project(args: dict) -> str:
    """更新项目元数据（template.meta.json）：更新描述或合并额外元数据字段"""
    project_name = args["projectName"]
    description = args.get("description")
    meta = args.get("meta")
    ws = _workspace_dir or "workspace"
    project_dir = Path(ws) / project_name
    if not project_dir.exists():
        return json.dumps({
            "status": "error", "error": f"项目 '{project_name}' 不存在，无法更新",
        }, ensure_ascii=False)
    meta_path = project_dir / "template.meta.json"
    data = {}
    if meta_path.exists():
        try:
            data = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            data = {}
    if description is not None:
        data["description"] = description
    if isinstance(meta, dict):
        data.update(meta)
    meta_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return json.dumps({
        "status": "ok", "projectName": project_name,
        "message": f"项目 '{project_name}' 元数据已更新",
        "meta": data,
    }, ensure_ascii=False)


def _zip_project_dir(project_dir: Path, zip_path: Path):
    """将项目目录递归打包为 zip（跳过隐藏目录与 __pycache__）"""
    from zipfile import ZipFile, ZIP_DEFLATED
    with ZipFile(zip_path, "w", ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(project_dir):
            dirs[:] = [d for d in dirs if not d.startswith(".") and d != "__pycache__"]
            for fname in files:
                fpath = Path(root) / fname
                arcname = fpath.relative_to(project_dir)
                zf.write(fpath, arcname.as_posix())


async def _export_project(args: dict) -> str:
    """导出项目包（zip）：默认导出到 workspace/_exports/<项目名>.zip"""
    project_name = args["projectName"]
    target_dir = args.get("targetDir", "")
    ws = _workspace_dir or "workspace"
    project_dir = Path(ws) / project_name
    if not project_dir.exists():
        return json.dumps({
            "status": "error", "error": f"项目 '{project_name}' 不存在，无法导出",
        }, ensure_ascii=False)
    out_dir = Path(target_dir) if target_dir else Path(ws) / "_exports"
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        return json.dumps({"status": "error", "error": f"无法创建导出目录: {e}"}, ensure_ascii=False)
    zip_path = out_dir / f"{project_name}.zip"
    try:
        _zip_project_dir(project_dir, zip_path)
    except OSError as e:
        return json.dumps({"status": "error", "error": f"导出失败: {e}"}, ensure_ascii=False)
    return json.dumps({
        "status": "ok", "projectName": project_name, "zipPath": str(zip_path),
        "message": f"项目 '{project_name}' 已导出为 {zip_path.name}",
    }, ensure_ascii=False)


async def _import_project(args: dict) -> str:
    """导入项目包（zip）为项目；内置 zip-slip 路径穿越防护"""
    project_name = args["projectName"]
    zip_path = args["zipPath"]
    ws = _workspace_dir or "workspace"
    src_zip = Path(zip_path)
    if not src_zip.exists():
        return json.dumps({
            "status": "error", "error": f"项目包 '{zip_path}' 不存在，无法导入",
        }, ensure_ascii=False)
    target_dir = Path(ws) / project_name
    if target_dir.exists():
        return json.dumps({
            "status": "error", "error": f"项目 '{project_name}' 已存在，请更换名称或先删除",
        }, ensure_ascii=False)
    from zipfile import ZipFile
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
        with ZipFile(src_zip) as zf:
            for info in zf.infolist():
                name = info.filename
                # zip-slip 防护：拒绝绝对路径与 .. 穿越
                if name.startswith("/") or ".." in name.replace("\\", "/").split("/"):
                    raise RuntimeError(f"项目包包含非法路径: {name}")
                dest = target_dir / name
                if info.is_dir():
                    dest.mkdir(parents=True, exist_ok=True)
                    continue
                dest.parent.mkdir(parents=True, exist_ok=True)
                dest.write_bytes(zf.read(info))
    except RuntimeError as e:
        return json.dumps({"status": "error", "error": str(e)}, ensure_ascii=False)
    except OSError as e:
        return json.dumps({"status": "error", "error": f"导入失败: {e}"}, ensure_ascii=False)
    return json.dumps({
        "status": "ok", "projectName": project_name,
        "message": f"项目包已导入为项目 '{project_name}'",
    }, ensure_ascii=False)


async def _create_from_template(args: dict) -> str:
    """基于模板中心（example 目录）的模板创建新项目"""
    project_name = args["projectName"]
    template_name = args["templateName"]
    return await _run_python_cli(["project", "create", project_name, "--template", template_name])


async def _preview_template(args: dict) -> str:
    """预览项目内模板的渲染结果（不写入文件）"""
    project_name = args["projectName"]
    template_path = args["templatePath"]
    return await _run_python_cli(["template", "preview", project_name, template_path])


# ====== 4.3 F3-3：技能库工具（技能可被 AI 工具调用）======

async def _list_skills(args: dict) -> str:
    from ai_hub.skills.engine import get_skills_engine
    skills = get_skills_engine().list_skills()
    return json.dumps({"status": "ok", "skills": skills, "total": len(skills)}, ensure_ascii=False)


async def _get_skill(args: dict) -> str:
    from ai_hub.skills.engine import get_skills_engine
    skill = get_skills_engine().get_skill(args["skillName"])
    if skill is None:
        return json.dumps({"status": "error", "error": f"技能 '{args['skillName']}' 不存在"}, ensure_ascii=False)
    return json.dumps({"status": "ok", "skill": skill}, ensure_ascii=False)


async def _enable_skill(args: dict) -> str:
    from ai_hub.skills.engine import get_skills_engine
    ok = get_skills_engine().enable_skill(args["skillName"])
    if not ok:
        return json.dumps({"status": "error", "error": f"技能 '{args['skillName']}' 不存在"}, ensure_ascii=False)
    return json.dumps({
        "status": "ok", "skillName": args["skillName"], "enabled": True,
        "message": f"技能 '{args['skillName']}' 已启用",
    }, ensure_ascii=False)


async def _disable_skill(args: dict) -> str:
    from ai_hub.skills.engine import get_skills_engine
    ok = get_skills_engine().disable_skill(args["skillName"])
    if not ok:
        return json.dumps({"status": "error", "error": f"技能 '{args['skillName']}' 不存在"}, ensure_ascii=False)
    return json.dumps({
        "status": "ok", "skillName": args["skillName"], "enabled": False,
        "message": f"技能 '{args['skillName']}' 已禁用",
    }, ensure_ascii=False)


async def _update_skill(args: dict) -> str:
    from ai_hub.skills.engine import get_skills_engine
    skill = get_skills_engine().save_skill(args["skillName"], args["content"])
    return json.dumps({
        "status": "ok", "skillName": skill.name,
        "message": f"技能 '{skill.name}' 已更新",
    }, ensure_ascii=False)


def init_tools():
    """初始化所有 Agent Tools"""
    register_tool(
        "list_projects",
        "列出所有项目及其结构",
        {
            "type": "object",
            "properties": {},
            "required": [],
        },
        _list_projects,
    )

    register_tool(
        "create_project",
        "⚠️ 仅用于从 example 目录复制模板项目。从模板中心创建项目请使用 create_project_intelligent。指定项目名称和可选的模板来源",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
                "templateName": {"type": "string", "description": "模板名称（可选）"},
            },
            "required": ["projectName"],
        },
        _create_project,
    )

    register_tool(
        "create_template",
        "将现有项目保存为可复用模板",
        {
            "type": "object",
            "properties": {
                "sourceProject": {"type": "string", "description": "源项目名称"},
                "templateName": {"type": "string", "description": "新模板名称"},
            },
            "required": ["sourceProject", "templateName"],
        },
        _create_template,
    )

    register_tool(
        "template_list",
        "列出所有可用的示例模板（模板中心）",
        {
            "type": "object",
            "properties": {},
            "required": [],
        },
        _list_templates,
    )

    register_tool(
        "template_delete",
        "删除示例模板（不可恢复，请谨慎使用）",
        {
            "type": "object",
            "properties": {
                "templateName": {"type": "string", "description": "模板名称"},
            },
            "required": ["templateName"],
        },
        _delete_template,
    )

    register_tool(
        "update_template",
        "修改模板文件内容",
        {
            "type": "object",
            "properties": {
                "templateName": {"type": "string", "description": "模板名称"},
                "filePath": {"type": "string", "description": "文件相对路径"},
                "content": {"type": "string", "description": "新内容"},
            },
            "required": ["templateName", "filePath", "content"],
        },
        _update_template,
    )

    register_tool(
        "render_config",
        "执行配置渲染，生成设备配置文件",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
            },
            "required": ["projectName"],
        },
        _render_config,
    )

    register_tool(
        "dry_run",
        "预演渲染，预览结果但不写入文件",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
            },
            "required": ["projectName"],
        },
        _dry_run,
    )

    register_tool(
        "validate_template",
        "校验 Jinja2 模板语法",
        {
            "type": "object",
            "properties": {
                "templateName": {"type": "string", "description": "模板名称"},
            },
            "required": ["templateName"],
        },
        _validate_template,
    )

    register_tool(
        "validate_excel",
        "校验 Excel 参数文件",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
            },
            "required": ["projectName"],
        },
        _validate_excel,
    )

    register_tool(
        "diff_compare",
        "对比渲染结果与已有输出差异",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
            },
            "required": ["projectName"],
        },
        _diff_compare,
    )

    register_tool(
        "read_file",
        "读取项目中的文件内容",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
                "filePath": {"type": "string", "description": "文件相对路径"},
            },
            "required": ["projectName", "filePath"],
        },
        _read_file,
    )

    register_tool(
        "search_files",
        "按名称或内容搜索项目文件",
        {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "搜索关键词"},
                "projectName": {"type": "string", "description": "限定项目（可选）"},
            },
            "required": ["query"],
        },
        _search_files,
    )

    register_tool(
        "create_project_intelligent",
        "智能创建项目：根据设备类型（switch/router/firewall）和厂商（huawei/cisco/h3c）自动生成 Jinja2 模板和 Excel 参数表。创建项目后可用 update_template 微调模板内容",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称（英文，如 my_switch_project）"},
                "deviceType": {"type": "string", "description": "设备类型：switch（交换机）、router（路由器）、firewall（防火墙）", "enum": ["switch", "router", "firewall"]},
                "vendor": {"type": "string", "description": "厂商：huawei（华为）、cisco（思科）、h3c（H3C）", "enum": ["huawei", "cisco", "h3c"]},
                "configDescription": {"type": "string", "description": "配置需求描述（可选，用于自定义模板）"},
            },
            "required": ["projectName", "deviceType", "vendor"],
        },
        _create_project_intelligent,
    )

    register_tool(
        "reverse_engineer_config",
        "从已有网络设备配置文本反推 Jinja2 模板和 Excel 参数表。粘贴完整的设备配置（如 show run 输出），自动识别并提取变量（IP、主机名、VLAN、SNMP、AAA 等），生成模板和参数文件",
        {
            "type": "object",
            "properties": {
                "configText": {"type": "string", "description": "完整的设备配置文本（如 show running-config 或 display current-configuration 的输出）"},
                "projectName": {"type": "string", "description": "新项目名称"},
                "vendor": {"type": "string", "description": "厂商：huawei/cisco/h3c", "enum": ["huawei", "cisco", "h3c"]},
                "deviceType": {"type": "string", "description": "设备类型：switch/router/firewall", "enum": ["switch", "router", "firewall"]},
            },
            "required": ["configText", "projectName"],
        },
        _reverse_engineer_config,
    )

    register_tool(
        "recommend_template",
        "根据设备类型和厂商推荐合适的配置模板。可以查询所有可用模板，也可以根据项目名称分析现有模板并给出优化建议。支持华为/思科/H3C 的交换机/路由器/防火墙模板",
        {
            "type": "object",
            "properties": {
                "deviceType": {"type": "string", "description": "设备类型：switch/router/firewall（可选）", "enum": ["switch", "router", "firewall"]},
                "vendor": {"type": "string", "description": "厂商：huawei/cisco/h3c（可选）", "enum": ["huawei", "cisco", "h3c"]},
                "projectName": {"type": "string", "description": "项目名称（可选，用于分析现有模板）"},
            },
            "required": [],
        },
        _recommend_template,
    )

    # ====== 新增工具注册 ======

    register_tool(
        "delete_project",
        "删除指定项目及其所有文件（不可恢复，请谨慎使用）",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
            },
            "required": ["projectName"],
        },
        _delete_project,
    )

    register_tool(
        "get_project_info",
        "获取项目详细信息：目录结构、文件列表、各子目录是否存在",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
            },
            "required": ["projectName"],
        },
        _get_project_info,
    )

    register_tool(
        "render_yaml",
        "渲染项目的 YAML 文件",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
            },
            "required": ["projectName"],
        },
        _render_yaml,
    )

    register_tool(
        "undo_render",
        "撤销最近一次渲染，恢复备份的输出文件",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
            },
            "required": ["projectName"],
        },
        _undo_render,
    )

    register_tool(
        "generate_labels",
        "生成 Word 格式的设备标签文件",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
            },
            "required": ["projectName"],
        },
        _generate_labels,
    )

    register_tool(
        "generate_label_md",
        "生成 Markdown 格式的设备标签文件，可在程序内直接查看",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
            },
            "required": ["projectName"],
        },
        _generate_label_md,
    )

    register_tool(
        "delete_labels",
        "删除项目标签文件，projectName 可传 \"all\" 清空所有项目",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称或 \"all\"（所有项目）"},
            },
            "required": ["projectName"],
        },
        _delete_labels,
    )

    register_tool(
        "delete_files",
        "删除项目输出文件。projectName 可传 \"all\" 清空所有项目。fileType: output（设备配置）、yaml（YAML）、output-sn、yaml-sn。默认 output",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称或 \"all\"（所有项目）"},
                "fileType": {"type": "string", "description": "文件类型",
                             "enum": ["output", "yaml", "output-sn", "yaml-sn"]},
            },
            "required": ["projectName"],
        },
        _delete_files,
    )

    register_tool(
        "list_project_files",
        "列出项目目录下的所有文件和子目录结构",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
            },
            "required": ["projectName"],
        },
        _list_project_files,
    )

    register_tool(
        "read_excel",
        "读取项目中的 Excel 文件内容（指定工作表），返回表头和数据行",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
                "fileName": {"type": "string", "description": "Excel 文件名（如 parameter.xlsx）"},
                "sheetName": {"type": "string", "description": "工作表名称（可选，默认第一个）"},
            },
            "required": ["projectName", "fileName"],
        },
        _read_excel,
    )

    register_tool(
        "write_excel",
        "向项目中的 Excel 文件写入数据",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
                "fileName": {"type": "string", "description": "Excel 文件名"},
                "data": {"type": "object", "description": "写入数据: {sheet, headers, rows}"},
            },
            "required": ["projectName", "fileName", "data"],
        },
        _write_excel,
    )

    register_tool(
        "write_text_file",
        "在项目中创建或覆盖文本文件（模板、配置等）",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
                "filePath": {"type": "string", "description": "相对于项目根目录的文件路径"},
                "content": {"type": "string", "description": "文件内容"},
            },
            "required": ["projectName", "filePath", "content"],
        },
        _write_text_file,
    )

    register_tool(
        "analyze_project",
        "分析项目：检查模板复杂度、变量使用、Excel 数据质量、模板与参数表的交叉引用，生成优化建议报告",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
            },
            "required": ["projectName"],
        },
        _analyze_project,
    )

    # ====== 4.3 F3-4：项目/模板操作工具补齐 ======

    register_tool(
        "update_project",
        "更新项目元数据（template.meta.json）：更新项目描述，或通过 meta 合并额外元数据字段（如 deviceType/tags）",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
                "description": {"type": "string", "description": "项目描述（可选）"},
                "meta": {"type": "object", "description": "额外元数据字段（可选，如 deviceType/tags）"},
            },
            "required": ["projectName"],
        },
        _update_project,
    )

    register_tool(
        "export_project",
        "导出项目包（zip）：默认导出到 workspace/_exports/<项目名>.zip，可用 targetDir 指定导出目录",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
                "targetDir": {"type": "string", "description": "导出目录（可选，默认 workspace/_exports）"},
            },
            "required": ["projectName"],
        },
        _export_project,
    )

    register_tool(
        "import_project",
        "导入项目包（zip）为项目：解压到 workspace/<项目名>（内置 zip-slip 路径穿越防护）",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "新项目名称"},
                "zipPath": {"type": "string", "description": "项目包 zip 文件路径"},
            },
            "required": ["projectName", "zipPath"],
        },
        _import_project,
    )

    register_tool(
        "create_from_template",
        "基于模板中心（example 目录）的模板创建新项目，指定项目名称与模板名称",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "新项目名称"},
                "templateName": {"type": "string", "description": "模板中心模板名称（如 example1/example2）"},
            },
            "required": ["projectName", "templateName"],
        },
        _create_from_template,
    )

    register_tool(
        "preview_template",
        "预览项目内模板的渲染结果（不写入文件）：指定项目名称与模板相对路径（如 templates/ASW.j2）",
        {
            "type": "object",
            "properties": {
                "projectName": {"type": "string", "description": "项目名称"},
                "templatePath": {"type": "string", "description": "模板文件相对路径（如 templates/ASW.j2）"},
            },
            "required": ["projectName", "templatePath"],
        },
        _preview_template,
    )

    # ====== 4.3 F3-3：技能库工具 ======

    register_tool(
        "list_skills",
        "列出技能库中所有技能（名称/启用状态/使用统计）",
        {
            "type": "object",
            "properties": {},
            "required": [],
        },
        _list_skills,
    )

    register_tool(
        "get_skill",
        "获取单个技能详情（含内容），供参考复用或向用户展示",
        {
            "type": "object",
            "properties": {
                "skillName": {"type": "string", "description": "技能名称"},
            },
            "required": ["skillName"],
        },
        _get_skill,
    )

    register_tool(
        "enable_skill",
        "启用技能（恢复其进入 AI 上下文）",
        {
            "type": "object",
            "properties": {
                "skillName": {"type": "string", "description": "技能名称"},
            },
            "required": ["skillName"],
        },
        _enable_skill,
    )

    register_tool(
        "disable_skill",
        "禁用技能（不再进入 AI 上下文，保留文件不删除）",
        {
            "type": "object",
            "properties": {
                "skillName": {"type": "string", "description": "技能名称"},
            },
            "required": ["skillName"],
        },
        _disable_skill,
    )

    register_tool(
        "update_skill",
        "更新（创建或覆盖）技能内容，技能会立即进入 AI 上下文",
        {
            "type": "object",
            "properties": {
                "skillName": {"type": "string", "description": "技能名称"},
                "content": {"type": "string", "description": "技能完整内容（Markdown）"},
            },
            "required": ["skillName", "content"],
        },
        _update_skill,
    )

    logger.info(f"Initialized {len(_tools)} Agent tools")