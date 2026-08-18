"""
AIDC 单项目四表格生成器（P1.2，AIDC 程序优化 PRD FR-A）。

一个 64 台试点 = **一个 MC 项目**，含 4 个工作簿（多 sheet），覆盖四网全部设备：

```
{项目}/
├── para.xlsx               project_para 声明
├── excel/
│   ├── hostname.xlsx       ① 设备表（22 台）
│   ├── connection.xlsx     ② 互联关系表 / 终端连接表 / MLAG表 / VLAN网关表
│   ├── ipaddress.xlsx      ③ 环回地址表 / 互联地址表 / 网段规划表
│   └── parameter.xlsx      ④ 全局参数表
├── templates/{角色}.j2
└── template.meta.json
```

角色 → 模板：SPINE/LEAF/STO_SPINE/STO_LEAF/BIZ_AGG/BIZ_ACCESS/OOB_AGG/OOB_ACCESS。
"""

import os
import json
from datetime import datetime

import pandas as pd

from .resolver import IntentContext
from .project_aidc import _info_template, _PEER_AS, _PLANE_SHEET, ROLE_SCENARIO
from .planner.validate import validate_context

# 固定 xlsx 元数据时间戳（openpyxl/zipfile 默认写当前时间 → 破坏字节级幂等）
# openpyxl 保存时忽略 properties.modified 覆写，故在 ZIP/XML 层直接改写。
_FIXED_TS = datetime(2026, 1, 1, 0, 0, 0)


def _fix_workbook_byte_idempotent(filepath: str):
    """重打包 xlsx：改写 core.xml created/modified + 固定 ZIP 条目 DOS 时间戳 → 字节级幂等。"""
    import io
    import re
    import zipfile
    ts = _FIXED_TS.strftime('%Y-%m-%dT%H:%M:%SZ')
    data = io.BytesIO()
    with zipfile.ZipFile(filepath, 'r') as zin:
        with zipfile.ZipFile(data, 'w', zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                content = zin.read(info.filename)
                if info.filename == 'docProps/core.xml':
                    xml = content.decode('utf-8', 'replace')
                    xml = re.sub(
                        r'(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)',
                        lambda m: f'{m.group(1)}{ts}{m.group(2)}', xml)
                    content = xml.encode('utf-8')
                new = zipfile.ZipInfo(info.filename,
                                      date_time=(_FIXED_TS.year, _FIXED_TS.month, _FIXED_TS.day, 0, 0, 0))
                new.compress_type = zipfile.ZIP_DEFLATED
                new.external_attr = info.external_attr
                zout.writestr(new, content)
    with open(filepath, 'wb') as fh:
        fh.write(data.getvalue())


def _write_sheet(df: pd.DataFrame, filepath: str, sheet_name: str, index: bool = False,
                 append: bool = False):
    """专业表格写入（FR-E）：表头加粗 + 冻结首行 + 自适应列宽 + 浅色表头。

    append=True 时追加到已存在工作簿（多 sheet 场景）。
    """
    if append:
        with pd.ExcelWriter(filepath, mode='a', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=index)
        _format_workbook(filepath, sheet_name)
        return
    df.to_excel(filepath, sheet_name=sheet_name, index=index, engine='openpyxl')
    _format_workbook(filepath, sheet_name)


def _format_workbook(filepath: str, sheet_name: str):
    """对指定 sheet 应用表头/冻结/列宽格式。"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    fill = PatternFill('solid', fgColor='DDEBF7')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    wb.save(filepath)
    ws = wb[sheet_name]
    # 表头样式
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill('solid', fgColor='DDEBF7')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
    # 冻结首行 + 自适应列宽
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    wb.save(filepath)


# 场景 -> 模板角色
_SCN_TO_ROLE = {
    'SPINE': 'SPINE', 'LEAF': 'LEAF', 'STO_SPINE': 'STO_SPINE', 'STO_LEAF': 'STO_LEAF',
    'BIZAGG': 'BIZ_AGG', 'BIZACC': 'BIZ_ACCESS', 'OOBAGG': 'OOB_AGG', 'OOBACC': 'OOB_ACCESS',
}
_SCN_PLANE = {
    'SPINE': '参数网', 'LEAF': '参数网', 'STO_SPINE': '存储网', 'STO_LEAF': '存储网',
    'BIZAGG': '业务&管理网', 'BIZACC': '业务&管理网', 'OOBAGG': '带外网', 'OOBACC': '带外网',
}


def _strip(val):
    return str(val).split('/')[0]


def _mask(prefix):
    from . import filters
    return filters.to_mask(f'0.0.0.0/{prefix}')


def _model_of(scn):
    role = _SCN_TO_ROLE[scn]
    return ROLE_SCENARIO[role][1]


class SingleProjectGenerator:
    """单项目四表格生成器（覆盖全部场景）。"""

    def __init__(self, ctx: IntentContext):
        self.ctx = ctx
        self.roles = set(_SCN_TO_ROLE[s] for s in ctx.device_params)  # 模板角色集合

    def _scenarios(self):
        return sorted(ctx for ctx in self.ctx.device_params)

    def _dev(self, scn, local, var_tail):
        return self.ctx.device_params.get(scn, {}).get(local, {}).get(f'{var_tail}{scn}{local}')

    def _list(self, scn, local, name):
        return self.ctx.lists.get(f'{scn}_{name}{local}', [])

    # ---- ① 设备表（H2：按四网拆 sheet；去 对端AS 列，MC-1） ----
    def build_device_table(self, plane=None):
        rows = []
        for scn in self._scenarios():
            if plane and _PLANE_SHEET.get(_SCN_PLANE[scn], _SCN_PLANE[scn]) != plane:
                continue
            for local, params in sorted(self.ctx.device_params[scn].items()):
                loopback = self._dev(scn, local, 'ipv4_LoopBack_P_')
                milo = self._dev(scn, local, 'ipv4_M-ILO_P_')
                asn = self._dev(scn, local, 'hostname_hostname_E_')
                lpre = str(loopback).split('/')[1] if loopback and '/' in str(loopback) else '32'
                mpre = str(milo).split('/')[1] if milo and '/' in str(milo) else '24'
                role = _SCN_TO_ROLE[scn]
                rows.append({
                    '设备名': self._dev(scn, local, 'hostname_hostname_B_'),
                    '型号': _model_of(scn),
                    '角色': role,
                    '环回接口': 'LoopBack0',
                    '环回IP': _strip(loopback),
                    '环回长度': int(lpre),
                    '管理接口': 'M-GigabitEthernet0/0/0',
                    '管理IP': _strip(milo),
                    '管理掩码': _mask(int(mpre)),
                    'BGP AS': asn if asn is not None else 65000,
                    'BGP多路径': self.ctx.globals.get('bgp_max_paths', 16),
                    'MLAG对': params.get('mlag_pair', ''),
                    'MLAG序号': params.get('mlag_system_number', ''),
                    'MLAG本端': params.get('mlag_keepalive', ''),
                    'MLAG对端': params.get('mlag_peer_keepalive', ''),
                    'SN': f'AIDC{scn}{local:03d}',
                })
        return pd.DataFrame(rows)

    # ---- ② connection.xlsx ----
    # 下联场景 -> 上联对端场景
    _PEER_SCN = {'LEAF': 'SPINE', 'STO_LEAF': 'STO_SPINE', 'BIZACC': 'BIZAGG', 'OOBACC': 'OOBAGG'}

    def _peer_hosts(self, scn):
        """取上联对端真实主机名列表（对端场景全部设备）。"""
        peer_scn = self._PEER_SCN[scn]
        hosts = []
        for local in sorted(self.ctx.device_params.get(peer_scn, {})):
            h = self._dev(peer_scn, local, 'hostname_hostname_B_')
            if h:
                hosts.append(h)
        return hosts

    def build_conn_table(self, plane=None):
        """IP规划地址表（对称表，H2：按四网拆 sheet；每链路行含 己端AS/对端AS，D-6）。

        LEAF/STO_LEAF/BIZ_ACCESS/OOBACC 方向生成（对端=真实 SPINE/AGG 主机名，避免伪设备）。
        列序 = 对称表 col_num=5：己端(设备/接口/IP/长度/AS) + 对端(设备/接口/IP/长度/AS)。
        对端接口 = 对端设备真实上联口（修复镜像时占位接口跨设备冲突导致的对端仅渲染到最后关联设备）。
        对端IP = 地址分配器产出的 ctx.bgp_peer_ip（同 /31、零冲突）。
        """
        rows = []
        for scn in self._scenarios():
            if not _is_leaf_direction(scn):
                continue
            if plane and _PLANE_SHEET.get(_SCN_PLANE[scn], _SCN_PLANE[scn]) != plane:
                continue
            peer_hosts = self._peer_hosts(scn)
            if not peer_hosts:
                continue
            for local, params in sorted(self.ctx.device_params[scn].items()):
                ips = self._list(scn, local, 'uplink_ip')
                peers = self._list(scn, local, 'bgp_peer_ip')
                pases = self._list(scn, local, 'bgp_peer_as')
                ports = self._list(scn, local, 'uplink_port')
                my_as = self._dev(scn, local, 'hostname_hostname_E_')
                # 每接入设备到同一对端的链路数（对端上联序列中该设备占用的槽位）
                per_peer = len(ports) // len(peer_hosts) if len(peer_hosts) else 0
                for idx, ip in enumerate(ips):
                    peer = peer_hosts[idx % len(peer_hosts)]
                    # 对端真实上联口：该链路在对端上联序列中的位置
                    #   k = (local-1) * per_peer + idx // len(peer_hosts)
                    #   （与 plan_builder._build_agg_uplinks 的 reverse 轮询顺序一致，镜像 key 唯一）
                    k = (local - 1) * per_peer + (idx // len(peer_hosts)) if per_peer else 0
                    rows.append({
                        '己端设备': self._dev(scn, local, 'hostname_hostname_B_'),
                        '己端接口': ports[idx] if idx < len(ports) else '',
                        '己端IP地址': ip,
                        '己端IP长度': 31,
                        '己端AS': my_as if my_as is not None else 65000,
                        '对端设备': peer,
                        '对端接口': self._peer_uplink_port(peer, k),
                        '对端IP地址': peers[idx] if idx < len(peers) else '',
                        '对端IP长度': 31,
                        '对端AS': pases[idx] if idx < len(pases) else '',
                        '备注信息': f'{_SCN_PLANE[scn]}上联',
                    })
        return pd.DataFrame(rows)

    def _peer_uplink_port(self, peer, k):
        """对端设备真实上联口：按对端主机名定位 (scn, local)，取其上联序列第 k 个口。"""
        for scn in self._scenarios():
            for local in sorted(self.ctx.device_params.get(scn, {})):
                if self._dev(scn, local, 'hostname_hostname_B_') == peer:
                    ports = self._list(scn, local, 'uplink_port')
                    return ports[k] if 0 <= k < len(ports) else ''
        return ''

    def build_terminal_table(self, plane=None):
        """终端连接表（H2：每接口一行，去逗号拼接；按四网拆 sheet）。"""
        rows = []
        for scn in self._scenarios():
            if scn in ('SPINE', 'STO_SPINE', 'BIZAGG', 'OOBAGG'):
                continue
            if plane and _PLANE_SHEET.get(_SCN_PLANE[scn], _SCN_PLANE[scn]) != plane:
                continue
            for local in sorted(self.ctx.device_params[scn]):
                ports = self._list(scn, local, _terminal_port_key(scn))
                vlans = self._list(scn, local, _terminal_vlan_key(scn))
                descs = self._list(scn, local, _terminal_desc_key(scn))
                if not ports:
                    continue
                host = self._dev(scn, local, 'hostname_hostname_B_')
                for i, p in enumerate(ports):
                    rows.append({
                        '己端设备': host,
                        '己端接口': p,
                        '己端VLAN': vlans[i] if i < len(vlans) else '',
                        '己端描述': descs[i] if i < len(descs) else '',
                        '备注信息': f'{_SCN_PLANE[scn]}下联',
                    })
        return pd.DataFrame(rows)

    def build_mlag_table(self):
        """MLAG 表（赋值表）：BIZ_ACCESS 成对。"""
        rows = []
        for local in sorted(self.ctx.device_params.get('BIZACC', {})):
            p = self.ctx.device_params['BIZACC'][local]
            if 'mlag_pair' not in p:
                continue
            rows.append({
                '己端设备': p.get('hostname_hostname_B_BIZACC%d' % local, ''),
                'MLAG对': p['mlag_pair'],
                'MLAG序号': p.get('mlag_system_number', ''),
                'MLAG本端': p.get('mlag_keepalive', ''),
                'MLAG对端': p.get('mlag_peer_keepalive', ''),
            })
        return pd.DataFrame(rows)

    def build_vlan_gw_table(self, plane=None):
        """VLAN 网关表（H2：每 VLAN 一行，去逗号拼接；按四网拆 sheet，带外无网关）。"""
        rows = []
        for scn in ('LEAF', 'STO_LEAF', 'BIZACC'):
            if plane and _PLANE_SHEET.get(_SCN_PLANE[scn], _SCN_PLANE[scn]) != plane:
                continue
            for local in sorted(self.ctx.device_params.get(scn, {})):
                vids = self._list(scn, local, 'vlan_id')
                gws = self._list(scn, local, 'vlan_gw')
                if not vids:
                    continue
                host = self._dev(scn, local, 'hostname_hostname_B_')
                for i, v in enumerate(vids):
                    rows.append({
                        '己端设备': host,
                        '网关VLAN': v,
                        '网关IP': gws[i] if i < len(gws) else '',
                        '备注信息': f'{_SCN_PLANE[scn]}网关',
                    })
        return pd.DataFrame(rows)

    # ---- ③ ipaddress.xlsx ----
    def build_loopback_table(self):
        rows = []
        for scn in self._scenarios():
            for local in sorted(self.ctx.device_params[scn]):
                rows.append({
                    '己端设备': self._dev(scn, local, 'hostname_hostname_B_'),
                    '环回接口': 'LoopBack0',
                    '环回IP': _strip(self._dev(scn, local, 'ipv4_LoopBack_P_')),
                    '环回长度': 32,
                })
        return pd.DataFrame(rows)

    def build_subnet_table(self):
        """网段规划表（参数表）：各平面网段。"""
        return pd.DataFrame([
            {'网段用途': '环回', '网段': '10.1.0.0/20', '掩码': '255.255.0.0'},
            {'网段用途': '计算网关', '网段': '10.1.16.0/20', '掩码': '255.255.240.0'},
            {'网段用途': '存储网关', '网段': '10.1.32.0/20', '掩码': '255.255.240.0'},
            {'网段用途': '业务网关', '网段': '10.1.48.0/20', '掩码': '255.255.240.0'},
            {'网段用途': '带外/管理', '网段': '10.1.64.0/21', '掩码': '255.255.248.0'},
            {'网段用途': '互联', '网段': '10.1.72.0/21', '掩码': '255.255.248.0'},
        ])

    # ---- ④ parameter.xlsx ----
    def build_param_table(self):
        from .project_aidc import AidcProjectGenerator
        # 复用 project_aidc 的参数表构建（含 PFC/CNP/AAA/NTP 等）
        return AidcProjectGenerator(self.ctx, {}, '全部').build_param_table()

    # ---- 写项目 ----
    def write(self, project_dir: str):
        os.makedirs(os.path.join(project_dir, 'excel'), exist_ok=True)
        os.makedirs(os.path.join(project_dir, 'templates'), exist_ok=True)

        excel = os.path.join(project_dir, 'excel')
        PLANES = ('参数网', '存储网', '业务网', '带外网')
        # H2（MC-2~5）：先构建全部平面 sheet（复用，避免重复计算）
        dev_dfs = {p: self.build_device_table(p) for p in PLANES}
        term_dfs = {p: self.build_terminal_table(p) for p in PLANES}
        gw_dfs = {p: self.build_vlan_gw_table(p) for p in ('参数网', '存储网', '业务网')}
        conn_dfs = {p: self.build_conn_table(p) for p in PLANES}

        # ① 设备表（四网拆 sheet，去对端AS）
        host_path = os.path.join(excel, 'hostname.xlsx')
        first = True
        for p in PLANES:
            if dev_dfs[p].empty:
                continue
            _write_sheet(dev_dfs[p], host_path, f'设备表-{p}', append=not first)
            first = False
        # ② connection.xlsx：终端连接表（每接口一行）+ VLAN网关表（每 VLAN 一行）
        conn_path = os.path.join(excel, 'connection.xlsx')
        first = True
        for p in PLANES:
            if term_dfs[p].empty:
                continue
            _write_sheet(term_dfs[p], conn_path, f'终端连接表-{p}', append=not first)
            first = False
        for p in ('参数网', '存储网', '业务网'):
            if gw_dfs[p].empty:
                continue
            _write_sheet(gw_dfs[p], conn_path, f'VLAN网关表-{p}', append=True)
        # ③ ipaddress.xlsx：IP规划地址表（对称表，四网拆）+ 环回/网段
        ip_path = os.path.join(excel, 'ipaddress.xlsx')
        first = True
        for p in PLANES:
            if conn_dfs[p].empty:
                continue
            _write_sheet(conn_dfs[p], ip_path, f'IP规划地址表-{p}', append=not first)
            first = False
        _write_sheet(self.build_loopback_table(), ip_path, '环回地址表', append=True)
        _write_sheet(self.build_subnet_table(), ip_path, '网段规划表', append=True)
        # ④ parameter.xlsx
        _write_sheet(self.build_param_table(), os.path.join(excel, 'parameter.xlsx'), '参数表')

        # para.xlsx 声明（H2：全部 sheet）
        para_rows = []
        for p in PLANES:
            if not dev_dfs[p].empty:
                para_rows.append(['hostname.xlsx', f'设备表-{p}', '赋值表', 0, 1])
        for p in PLANES:
            if not term_dfs[p].empty:
                para_rows.append(['connection.xlsx', f'终端连接表-{p}', '赋值表', 0, 2])
        for p in ('参数网', '存储网', '业务网'):
            if not gw_dfs[p].empty:
                para_rows.append(['connection.xlsx', f'VLAN网关表-{p}', '赋值表', 0, 2])
        for p in PLANES:
            if not conn_dfs[p].empty:
                para_rows.append(['ipaddress.xlsx', f'IP规划地址表-{p}', '对称表', 5, 2])
        para_rows += [['ipaddress.xlsx', '环回地址表', '赋值表', 0, 2],
                      ['ipaddress.xlsx', '网段规划表', '参数表', 0, 1],
                      ['parameter.xlsx', '参数表', '参数表', 0, 1]]
        proj_para = pd.DataFrame(para_rows,
                                 columns=['工作簿名称', '工作表名称', '工作表类型', '对称列数', 'key列数'])
        proj_para.to_excel(os.path.join(project_dir, 'para.xlsx'), index=False, sheet_name='project_para')

        # 固定全部 xlsx 时间戳（字节级幂等：同 plan → 同文件）
        for f in [os.path.join(project_dir, 'para.xlsx')] + \
                 [os.path.join(excel, n) for n in ('hostname.xlsx', 'connection.xlsx', 'ipaddress.xlsx', 'parameter.xlsx')]:
            _fix_workbook_byte_idempotent(f)

        # 每角色模板
        for role in self.roles:
            plane = _SCN_PLANE[_role_to_scn(role)]
            with open(os.path.join(project_dir, 'templates', f'{role}.j2'), 'w', encoding='utf-8') as f:
                f.write(_info_template(role, plane))

        issues = validate_context(self.ctx)
        meta = {
            'name': os.path.basename(project_dir.rstrip('/')),
            'source': 'intent-project-single-gen',
            'license': 'private',
            'planes': sorted(set(_SCN_PLANE.values())),
            'roles': sorted(self.roles),
            'tunables': ['PFC队列', 'CNP队列'],
            'generator': 'intent.project_single.SingleProjectGenerator',
            'version': '0.3',
            'validation': {'ok': len(issues) == 0, 'issue_count': len(issues), 'issues': issues[:20]},
        }
        # G3.2 + 契约 v1.2（M-1）：桥接标识与项目身份透传（AL plan → MC 项目，判别规则见契约 §1.4/§6.2）
        bridge = getattr(self.ctx, 'bridge', None)
        if bridge:
            meta['source'] = bridge.get('source', 'autolink')
            meta['projectType'] = bridge.get('projectType', 'aidc')
            meta['bridgeVersion'] = bridge.get('bridgeVersion', '1.0')
            meta['originPlan'] = bridge.get('originPlan', '')
            for k in ('originProjectId', 'originProjectName', 'originSite', 'originPlanVersion', 'planHash'):
                if bridge.get(k):
                    meta[k] = bridge.get(k)
        with open(os.path.join(project_dir, 'template.meta.json'), 'w', encoding='utf-8') as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)
        with open(os.path.join(project_dir, 'README.md'), 'w', encoding='utf-8') as f:
            f.write(f'# {meta["name"]}\n\nAIDC 64 台试点单项目（四表格多 sheet，四网合一）。'
                    f'可调参数：PFC队列/CNP队列（0-7）。\n')
        return project_dir


def _is_leaf_direction(scn):
    return scn in ('LEAF', 'STO_LEAF', 'BIZACC', 'OOBACC')


def _terminal_port_key(scn):
    return {'LEAF': 'gpu_port', 'STO_LEAF': 'gpu_port',
            'BIZACC': 'biz_port', 'OOBACC': 'downlink_port'}[scn]


def _terminal_vlan_key(scn):
    return {'LEAF': 'gpu_vlan', 'STO_LEAF': 'gpu_vlan',
            'BIZACC': 'biz_vlan', 'OOBACC': 'downlink_vlan'}[scn]


def _terminal_desc_key(scn):
    return {'LEAF': 'gpu_desc', 'STO_LEAF': 'gpu_desc',
            'BIZACC': 'biz_desc', 'OOBACC': 'downlink_desc'}[scn]


def _role_to_scn(role):
    inv = {v: k for k, v in _SCN_TO_ROLE.items()}
    return inv.get(role, role)


def generate_single_pilot64_project(project_dir: str, ctx: IntentContext) -> str:
    """生成单项目（四表格多 sheet，四网合一）。"""
    return SingleProjectGenerator(ctx).write(project_dir)
