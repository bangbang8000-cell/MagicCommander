import json
import logging
import warnings
from base import Base
from ExcelToLabel import exceltolabel, exceltomarkdown
import os
import shutil
import pandas as pd
from pandas import read_excel
from time import strftime, localtime
from datetime import datetime
import hashlib
from config import WORKSPACE_DIR

logger = logging.getLogger(__name__)


class PreProcessing:

    def __init__(self):
        self.workspace = WORKSPACE_DIR
        self.project_para = []
        self.project_name = []
        self.target_project_name = []
        self.target_project_num = []

    def _get_base_path(self):
        """获取工作区目录"""
        return self.workspace

    def _calc_render_steps(self, include_template_render: bool = True):
        """按项目表数量计算真实渲染步骤数。"""
        extra_steps = 2 if include_template_render else 1
        total = 0
        for i, name in enumerate(self.target_project_name):
            para = self.project_para[self.target_project_num[i]].get(name, [])
            total += len(para) + extra_steps
        return max(total, 1)

    def _emit_progress(self, current_step: int, total_steps: int, message: str, **data):
        """输出统一进度事件。"""
        print(json.dumps({
            'status': 'progress',
            'message': message,
            'data': {
                'step': current_step,
                'totalSteps': total_steps,
                'progress': min(100, int(current_step / total_steps * 100)),
                **data,
            }
        }, ensure_ascii=False))

    def _safe_remove_dir(self, dir_path: str, success_msg: str, fail_msg: str):
        """安全删除目录（性能优化版）"""
        import time
        start_time = time.time()
        
        try:
            if not os.path.exists(dir_path):
                logger.warning(fail_msg)
                return False
                
            # 使用 shutil.rmtree 删除目录，并设置忽略错误
            shutil.rmtree(dir_path, ignore_errors=False, onerror=self._rmtree_error_handler)
            
            # 计算删除操作的执行时间
            execution_time = time.time() - start_time
            
            logger.info(f"{success_msg} (耗时: {execution_time:.2f}秒)")
            return True
            
        except PermissionError as info:
            logger.error(f"权限错误: {info}")
            return False
        except Exception as e:
            logger.error(f'删除异常，请核查！{e}', exc_info=True)
            return False
    
    def _rmtree_error_handler(self, func, path, exc_info):
        """处理 shutil.rmtree 的错误"""
        import stat
        
        try:
            # 尝试修改文件权限
            os.chmod(path, stat.S_IWUSR)
            # 再次尝试删除
            func(path)
        except Exception as e:
            logger.error(f"无法删除 {path}: {e}", exc_info=True)
            raise

    def read_MC_para(self, excel_MC: str):
        """读取项目名称，并合并工作区中已存在的项目目录。"""
        import time
        start_time = time.time()
        
        mid_path = self._get_base_path()
        filepath = os.path.join(mid_path, excel_MC)
        self.project_name = []
        
        need_sync_mc_para = False
        try:
            data = read_excel(filepath, sheet_name=None, keep_default_na=False)
            for sheet, value in data.items():
                if sheet == '项目名称':
                    for i in value.index.values:
                        project_name = str(value['项目名称'][i]).strip()
                        project_dir = os.path.join(mid_path, project_name)
                        has_project_shape = project_name and os.path.isdir(project_dir) and any(
                            os.path.exists(os.path.join(project_dir, name))
                            for name in ('para.xlsx', 'excel', 'templates', 'output', 'yaml')
                        )
                        if has_project_shape and project_name not in self.project_name:
                            self.project_name.append(project_name)
                        elif project_name:
                            need_sync_mc_para = True
        except Exception as e:
            need_sync_mc_para = True
            logger.error(f"读取项目名称失败: {e}", exc_info=True)

        try:
            if os.path.isdir(mid_path):
                ignored_dirs = {'__pycache__', 'assets', 'node_modules'}
                for entry in sorted(os.listdir(mid_path)):
                    project_dir = os.path.join(mid_path, entry)
                    if entry.startswith('.') or entry in ignored_dirs or not os.path.isdir(project_dir):
                        continue
                    has_project_shape = any(
                        os.path.exists(os.path.join(project_dir, name))
                        for name in ('para.xlsx', 'excel', 'templates', 'output', 'yaml')
                    )
                    if has_project_shape and entry not in self.project_name:
                        self.project_name.append(entry)
                        need_sync_mc_para = True
        except Exception as e:
            logger.error(f"扫描项目目录失败: {e}", exc_info=True)

        if need_sync_mc_para:
            try:
                pd.DataFrame({'项目名称': self.project_name}).to_excel(
                    filepath,
                    sheet_name='项目名称',
                    index=False,
                    header=True,
                )
                logger.info(f"已同步项目登记表: {filepath}")
            except Exception as e:
                logger.error(f"同步项目登记表失败: {e}", exc_info=True)

        execution_time = time.time() - start_time
        logger.info(f"读取项目名称完成 (耗时: {execution_time:.2f}秒, 项目数量: {len(self.project_name)})")

    def process_project_num(self, word: str):
        """处理项目编号"""
        if word == 'all':
            self.target_project_name = self.project_name[:]
            self.target_project_num = list(range(0, len(self.target_project_name)))
        else:
            num = word.split('/')
            for n in num:
                idx = int(n) - 1
                self.target_project_name.append(self.project_name[idx])
                self.target_project_num.append(idx)

    def read_project_para(self, name: str, excel_para: str, index: int | None = None):
        """实现各个项目具体的参数提取功能（性能优化版）"""
        import time
        start_time = time.time()
        project_list = []

        mid_path = self._get_base_path()
        filepath = os.path.join(mid_path, name, excel_para)

        try:
            data = read_excel(filepath, sheet_name=['project_para'], keep_default_na=False)
            for sheet, value in data.items():
                if sheet == 'project_para':
                    for i in value.index.values:
                        tmp_list = [
                            {'工作簿名称': str(value['工作簿名称'][i]).strip()},
                            {'工作表名称': str(value['工作表名称'][i]).strip()},
                            {'工作表类型': str(value['工作表类型'][i]).strip()},
                            {'对称列数': value['对称列数'][i]},
                            {'key列数': value['key列数'][i]},
                        ]
                        project_list.append(tmp_list)

            if index is None:
                self.project_para.append({name: project_list})
            else:
                # 保证列表长度后按索引写入，支持只读取目标项目
                while len(self.project_para) <= index:
                    self.project_para.append(None)
                self.project_para[index] = {name: project_list}

            # 计算读取操作的执行时间
            execution_time = time.time() - start_time
            logger.info(f"读取项目 '{name}' 参数完成 (耗时: {execution_time:.2f}秒, 参数数量: {len(project_list)})")

        except Exception as e:
            logger.error(f"读取项目 '{name}' 参数失败: {e}", exc_info=True)

    def _load_target_para(self):
        """仅为目标项目读取 para.xlsx（性能优化：不再全量读取所有项目）"""
        self.project_para = [None] * len(self.project_name)
        for idx in self.target_project_num:
            self.read_project_para(self.project_name[idx], 'para.xlsx', index=idx)

    def _process_project_sheets(self, ba, name: str, project_idx: int, current_step: int, total_steps: int) -> int:
        """读取并分发项目的一张张参数表（赋值表/对称表/参数表），返回更新后的 current_step。"""
        para = self.project_para[project_idx][name]
        for p in para:
            project_excel_name = p[0]['工作簿名称']
            project_sheet_name = p[1]['工作表名称']
            project_sheet_type = p[2]['工作表类型']
            project_sheet_assign_num = int(str(p[3]['对称列数']).strip())
            project_sheet_key_num = int(str(p[4]['key列数']).strip())

            if project_sheet_type == '赋值表':
                ba.read_assign_table(project_excel_name, project_sheet_name, 'excel', name, project_sheet_key_num)
            elif project_sheet_type == '对称表':
                ba.read_symmetrice_table(project_excel_name, project_sheet_name, 'excel', name, project_sheet_assign_num, project_sheet_key_num)
            elif project_sheet_type == '参数表':
                ba.read_para(project_excel_name, project_sheet_name, 'excel', name)
            else:
                continue

            current_step += 1
            self._emit_progress(
                current_step,
                total_steps,
                f'完成{project_excel_name} {project_sheet_name} 的数据提取',
                project=name,
                excel=project_excel_name,
                sheet=project_sheet_name,
                type=project_sheet_type,
                action='extract_sheet',
            )
        return current_step

    def _backup_output(self, project_dir: str) -> str | None:
        """渲染前备份 output 目录，支持撤销恢复。保留最近 MAX_OUTPUT_BACKUPS 份，超出自动轮转。"""
        output_dir = os.path.join(project_dir, 'output')
        if not os.path.exists(output_dir) or not os.listdir(output_dir):
            return None

        backup_dir = os.path.join(project_dir, '.output_backups')
        os.makedirs(backup_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = os.path.join(backup_dir, timestamp)
        shutil.copytree(output_dir, backup_path)
        logger.info(f'已备份 output 到 {backup_path}')

        # 轮转：仅保留最近 MAX_OUTPUT_BACKUPS 份备份，防止磁盘无限膨胀
        MAX_OUTPUT_BACKUPS = 5
        backups = sorted(os.listdir(backup_dir), reverse=True)
        for old in backups[MAX_OUTPUT_BACKUPS:]:
            try:
                shutil.rmtree(os.path.join(backup_dir, old))
                logger.info(f'已轮转删除旧备份: {old}')
            except Exception as e:
                logger.warning(f'轮转删除备份失败 {old}: {e}')

        return backup_path

    def _restore_backup(self, project_dir: str) -> bool:
        """恢复最近一次备份"""
        backup_dir = os.path.join(project_dir, '.output_backups')
        if not os.path.exists(backup_dir):
            return False
        backups = sorted(os.listdir(backup_dir), reverse=True)
        if not backups:
            return False
        latest = os.path.join(backup_dir, backups[0])
        output_dir = os.path.join(project_dir, 'output')
        if os.path.exists(output_dir):
            shutil.rmtree(output_dir)
        shutil.copytree(latest, output_dir)
        logger.info(f'已恢复备份: {backups[0]}')
        return True

    def _compute_project_cache_key(self, project_dir: str) -> str:
        """计算项目输入指纹（para.xlsx + excel + 模板 的 mtime/size，避免全量读文件）。"""
        hasher = hashlib.sha256()
        inputs = [os.path.join(project_dir, 'para.xlsx')]
        excel_dir = os.path.join(project_dir, 'excel')
        template_dir = os.path.join(project_dir, 'templates')
        if os.path.isdir(excel_dir):
            inputs.extend(sorted(os.path.join(excel_dir, f) for f in os.listdir(excel_dir) if f.endswith(('.xlsx', '.xls'))))
        if os.path.isdir(template_dir):
            inputs.extend(sorted(os.path.join(template_dir, f) for f in os.listdir(template_dir) if f.endswith(('.j2', '.jinja'))))

        for path in inputs:
            if os.path.exists(path):
                st = os.stat(path)
                hasher.update(path.encode('utf-8'))
                hasher.update(str(st.st_size).encode())
                hasher.update(str(int(st.st_mtime)).encode())
        return hasher.hexdigest()[:16]

    def _get_project_cache(self, project_dir: str, cache_key: str) -> bool:
        """判断项目输入指纹是否命中缓存（命中则输出可复用）。"""
        cache_meta = os.path.join(project_dir, '.render_cache', 'project.key')
        try:
            if os.path.exists(cache_meta):
                with open(cache_meta, encoding='utf-8') as f:
                    return f.read().strip() == cache_key
        except Exception:
            pass
        return False

    def _save_project_cache(self, project_dir: str, cache_key: str):
        """保存项目输入指纹到缓存。"""
        cache_dir = os.path.join(project_dir, '.render_cache')
        try:
            os.makedirs(cache_dir, exist_ok=True)
            with open(os.path.join(cache_dir, 'project.key'), 'w', encoding='utf-8') as f:
                f.write(cache_key)
        except Exception as e:
            logger.warning(f'保存渲染缓存失败: {e}')

    @staticmethod
    def _latest_timestamp_dir(base_dir: str) -> str | None:
        """返回目录下最新的时间戳子目录名。"""
        if not os.path.isdir(base_dir):
            return None
        dirs = [d for d in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, d))]
        return max(dirs) if dirs else None

    def _reuse_latest_output(self, project_dir: str, time_str: str, out_name_type: str) -> bool:
        """缓存命中：将最近一次渲染输出复制到新时间戳目录，避免重复提取/渲染。"""
        output_name = 'output' if out_name_type == 'device_name' else 'output-sn'
        yaml_name = 'yaml' if out_name_type == 'device_name' else 'yaml-sn'
        output_base = os.path.join(project_dir, output_name)
        latest = self._latest_timestamp_dir(output_base)
        if not latest:
            return False
        try:
            shutil.copytree(os.path.join(output_base, latest), os.path.join(output_base, time_str))
            yaml_base = os.path.join(project_dir, yaml_name)
            latest_yaml = self._latest_timestamp_dir(yaml_base)
            if latest_yaml:
                shutil.copytree(os.path.join(yaml_base, latest_yaml), os.path.join(yaml_base, time_str))
            logger.info(f'项目输入未变化，已复用渲染结果: {os.path.basename(project_dir)}')
            return True
        except Exception as e:
            logger.warning(f'复用渲染结果失败，将重新渲染: {e}')
            return False

    def execute_render(self, word: str, out_name_type: str):
        """执行选中项目的渲染"""
        time_str = strftime("%Y_%m_%d_%H_%M_%S", localtime())
        self.target_project_name = []
        self.target_project_num = []
        self.project_para = []
        self.process_project_num(word)

        self._load_target_para()

        total_steps = self._calc_render_steps(include_template_render=True)
        current_step = 0

        for i in range(0, len(self.target_project_name)):
            name = self.target_project_name[i]
            project_dir = os.path.join(self.workspace, name)

            # 缓存快速路径：输入指纹未变化且已有输出 → 复用，跳过提取与渲染
            cache_key = self._compute_project_cache_key(project_dir)
            if self._get_project_cache(project_dir, cache_key) and self._reuse_latest_output(project_dir, time_str, out_name_type):
                current_step += 1
                self._emit_progress(
                    current_step,
                    total_steps,
                    f'项目 {name} 输入未变化，已复用上次渲染结果',
                    project=name,
                    action='cache_hit',
                )
                continue

            # 渲染前自动备份 output 目录（仅在真正重新渲染时）
            self._backup_output(project_dir)

            ba = Base(self.workspace)
            current_step = self._process_project_sheets(ba, name, self.target_project_num[i], current_step, total_steps)

            if out_name_type == 'device_name':
                ba.out_base_info('yaml', name, time_str, 'device_name')
                current_step += 1
                self._emit_progress(
                    current_step,
                    total_steps,
                    f'项目 {name} 完成yaml文件的保存',
                    project=name,
                    action='yaml_save',
                )

                ba.render_txt('templates', name, time_str, 'device_name')
                current_step += 1
                self._emit_progress(
                    current_step,
                    total_steps,
                    f'项目 {name} jinja2运行完毕',
                    project=name,
                    action='jinja_render',
                )
            elif out_name_type == 'device_sn':
                ba.out_base_info('yaml-sn', name, time_str, 'device_sn')
                current_step += 1
                self._emit_progress(
                    current_step,
                    total_steps,
                    f'项目 {name} 完成yaml文件的保存',
                    project=name,
                    action='yaml_save',
                )

                ba.render_txt('templates', name, time_str, 'device_sn')
                current_step += 1
                self._emit_progress(
                    current_step,
                    total_steps,
                    f'项目 {name} jinja2运行完毕',
                    project=name,
                    action='jinja_render',
                )

            self._save_project_cache(project_dir, cache_key)

        # 4.8.0（F8-5 / 48-e）：渲染批次交付清单 manifest.json（逐文件 name/size/sha256 + 统计 + render_hash）
        output_name = 'output' if out_name_type == 'device_name' else 'output-sn'
        for _name in self.target_project_name:
            try:
                from intent.delivery import write_batch_manifest
                write_batch_manifest(os.path.join(self.workspace, _name), output_name, time_str)
            except Exception as _e:
                logger.warning(f'生成渲染批次清单失败: {_name}: {_e}')

        print(json.dumps({
            'status': 'complete',
            'message': '程序运行结束，请在目标项目的output文件夹内查看输出结果！',
            'data': {
                'totalSteps': total_steps,
                'completedSteps': total_steps,
                'progress': 100
            }
        }, ensure_ascii=False))

    def execute_dry_run(self, word: str, out_name_type: str):
        """执行渲染预览：不写文件，仅返回渲染输出内容"""
        self.target_project_name = []
        self.target_project_num = []
        self.project_para = []
        self.process_project_num(word)

        self._load_target_para()

        total_steps = self._calc_render_steps(include_template_render=True)
        current_step = 0
        results = []

        for i in range(0, len(self.target_project_name)):
            name = self.target_project_name[i]
            ba = Base(self.workspace)
            current_step = self._process_project_sheets(ba, name, self.target_project_num[i], current_step, total_steps)

            # 只渲染不写文件
            device_results = ba.render_dry_run('templates', name, out_name_type)
            current_step += 1
            self._emit_progress(
                current_step,
                total_steps,
                f'项目 {name} 渲染预览完成（{len(device_results)} 个设备）',
                project=name,
                action='dry_run',
            )
            results.extend(device_results)

        print(json.dumps({
            'status': 'complete',
            'message': f'渲染预览完成，共 {len(results)} 个设备',
            'data': {
                'totalSteps': total_steps,
                'completedSteps': total_steps,
                'progress': 100,
                'results': results,
            }
        }, ensure_ascii=False))

    def execute_template_preview(self, word: str, template_file: str):
        """调试沙盒：仅渲染指定模板文件，对项目内每台设备输出预览（不写文件）"""
        self.target_project_name = []
        self.target_project_num = []
        self.project_para = []
        self.process_project_num(word)

        self._load_target_para()

        results = []
        for i in range(0, len(self.target_project_name)):
            name = self.target_project_name[i]
            ba = Base(self.workspace)
            self._process_project_sheets(ba, name, self.target_project_num[i], 0, 1)
            results.extend(ba.render_preview(template_file, name))

        print(json.dumps({
            'status': 'complete',
            'message': f'模板预览完成，共 {len(results)} 个设备',
            'data': {
                'results': results,
            }
        }, ensure_ascii=False))

    def validate_template(self, word: str):
        """校验 Jinja2 模板语法"""
        self.target_project_name = []
        self.target_project_num = []
        self.process_project_num(word)

        results = []
        for name in self.target_project_name:
            templates_dir = os.path.join(self.workspace, name, 'templates')
            if not os.path.isdir(templates_dir):
                results.append({
                    'project': name,
                    'status': 'warning',
                    'message': '模板目录不存在',
                    'errors': [],
                })
                continue

            errors = []
            try:
                from jinja2 import Environment, FileSystemLoader, TemplateSyntaxError
                env = Environment(loader=FileSystemLoader(templates_dir))
                for fname in sorted(os.listdir(templates_dir)):
                    if not fname.endswith('.j2'):
                        continue
                    fpath = os.path.join(templates_dir, fname)
                    try:
                        with open(fpath, 'r', encoding='utf-8') as f:
                            source = f.read()
                        env.parse(source)
                    except TemplateSyntaxError as e:
                        errors.append({
                            'file': fname,
                            'line': e.lineno,
                            'message': e.message,
                        })
                    except Exception as e:
                        errors.append({
                            'file': fname,
                            'line': 0,
                            'message': f'读取失败: {e}',
                        })
            except Exception as e:
                errors.append({
                    'file': '',
                    'line': 0,
                    'message': f'初始化校验环境失败: {e}',
                })

            results.append({
                'project': name,
                'status': 'pass' if len(errors) == 0 else 'fail',
                'message': f'校验完成: {len(errors)} 个错误' if errors else '所有模板语法正确',
                'errors': errors,
            })

        print(json.dumps({
            'status': 'success',
            'message': f'模板校验完成，共 {len(results)} 个项目',
            'data': {'results': results},
        }, ensure_ascii=False))

    def validate_excel(self, word: str):
        """校验 Excel 数据完整性"""
        self.target_project_name = []
        self.target_project_num = []
        self.process_project_num(word)

        results = []
        for name in self.target_project_name:
            project_dir = os.path.join(self.workspace, name)
            warnings = []

            para_path = os.path.join(project_dir, 'para.xlsx')
            if not os.path.exists(para_path):
                results.append({
                    'project': name,
                    'status': 'fail',
                    'message': '缺少 para.xlsx 参数文件',
                    'warnings': [],
                })
                continue

            try:
                df = pd.read_excel(para_path, sheet_name=None, keep_default_na=False)
                if 'project_para' not in df:
                    warnings.append({
                        'type': 'missing_sheet',
                        'message': 'para.xlsx 缺少 project_para 工作表',
                    })
                else:
                    sheet = df['project_para']
                    for idx, row in sheet.iterrows():
                        workbook = str(row.get('工作簿名称', '')).strip()
                        sheet_name_excel = str(row.get('工作表名称', '')).strip()
                        if workbook and sheet_name_excel:
                            excel_path = os.path.join(project_dir, 'excel', workbook)
                            if not os.path.exists(excel_path):
                                warnings.append({
                                    'type': 'missing_excel',
                                    'file': workbook,
                                    'message': f'Excel 文件不存在: {workbook}',
                                })
                            else:
                                try:
                                    edf = pd.read_excel(excel_path, sheet_name=None, keep_default_na=False)
                                    if sheet_name_excel not in edf:
                                        warnings.append({
                                            'type': 'missing_sheet',
                                            'file': workbook,
                                            'sheet': sheet_name_excel,
                                            'message': f'工作表 "{sheet_name_excel}" 在 {workbook} 中不存在',
                                        })
                                    elif len(edf[sheet_name_excel]) == 0:
                                        warnings.append({
                                            'type': 'empty_sheet',
                                            'file': workbook,
                                            'sheet': sheet_name_excel,
                                            'message': f'工作表 "{sheet_name_excel}" 在 {workbook} 中为空',
                                        })
                                except Exception as e:
                                    warnings.append({
                                        'type': 'read_error',
                                        'file': workbook,
                                        'message': f'无法读取 {workbook}: {e}',
                                    })

                has_templates = any(os.path.isdir(os.path.join(project_dir, d)) and d == 'templates'
                                    for d in (os.listdir(project_dir) if os.path.isdir(project_dir) else []))
                if not has_templates:
                    warnings.append({
                        'type': 'missing_dir',
                        'message': '缺少 templates 目录',
                    })

            except Exception as e:
                warnings.append({
                    'type': 'read_error',
                    'message': f'无法读取 para.xlsx: {e}',
                })

            results.append({
                'project': name,
                'status': 'pass' if len(warnings) == 0 else 'warn',
                'message': f'校验完成: {len(warnings)} 个警告' if warnings else '所有数据完整',
                'warnings': warnings,
            })

        print(json.dumps({
            'status': 'success',
            'message': f'Excel 数据校验完成，共 {len(results)} 个项目',
            'data': {'results': results},
        }, ensure_ascii=False))

    def execute_yaml(self, word: str, out_name_type: str = 'device_name'):
        """执行选中项目的yaml创建

        out_name_type: 'device_name' → yaml/ 目录，'device_sn' → yaml-sn/ 目录
        """
        time_str = strftime("%Y_%m_%d_%H_%M_%S", localtime())
        self.target_project_name = []
        self.target_project_num = []
        self.project_para = []
        self.process_project_num(word)

        self._load_target_para()

        total_steps = self._calc_render_steps(include_template_render=False)
        current_step = 0

        yaml_dir = 'yaml-sn' if out_name_type == 'device_sn' else 'yaml'

        for i in range(0, len(self.target_project_name)):
            name = self.target_project_name[i]
            ba = Base(self.workspace)
            current_step = self._process_project_sheets(ba, name, self.target_project_num[i], current_step, total_steps)

            ba.out_base_info(yaml_dir, name, time_str, out_name_type)
            current_step += 1
            self._emit_progress(
                current_step,
                total_steps,
                f'项目 {name} 完成yaml文件的保存',
                project=name,
                action='yaml_save',
            )

        print(json.dumps({
            'status': 'complete',
            'message': '程序运行结束，请在目标项目的yaml文件夹内查看输出结果！',
            'data': {
                'totalSteps': total_steps,
                'completedSteps': total_steps,
                'progress': 100
            }
        }, ensure_ascii=False))

    def _write_demo_excel(self, path: str, sheet_name: str, rows: list):
        """写入示例 Excel，并保持表头/列宽便于新手查看。"""
        df = pd.DataFrame(rows)
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter

            wb = load_workbook(path)
            ws = wb[sheet_name]
            header_fill = PatternFill('solid', fgColor='1F4E78')
            header_font = Font(color='FFFFFF', bold=True, name='Arial')
            thin = Side(style='thin', color='D9DEE7')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            zebra = [PatternFill('solid', fgColor='FFFFFF'), PatternFill('solid', fgColor='F7F9FC')]

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.font = Font(name='Arial')
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row_index in range(2, ws.max_row + 1):
                fill = zebra[(row_index - 2) % 2]
                for col_index in range(1, ws.max_column + 1):
                    ws.cell(row_index, col_index).fill = fill

            for col_index in range(1, ws.max_column + 1):
                max_len = max(
                    len(str(ws.cell(row=row_index, column=col_index).value or ''))
                    for row_index in range(1, ws.max_row + 1)
                )
                ws.column_dimensions[get_column_letter(col_index)].width = min(max(max_len + 2, 10), 45)

            wb.save(path)
        except Exception as e:
            logger.warning(f'示例 Excel 样式写入失败，不影响数据生成: {e}')

    def _create_demo_project_files(self, project_dir: str):
        """生成可直接渲染的接入交换机示例项目。"""
        excel_dir = os.path.join(project_dir, 'excel')
        templates_dir = os.path.join(project_dir, 'templates')
        os.makedirs(excel_dir, exist_ok=True)
        os.makedirs(templates_dir, exist_ok=True)
        os.makedirs(os.path.join(project_dir, 'output'), exist_ok=True)
        os.makedirs(os.path.join(project_dir, 'yaml'), exist_ok=True)

        self._write_demo_excel(os.path.join(project_dir, 'para.xlsx'), 'project_para', [
            {'工作簿名称': 'hostname.xlsx', '工作表名称': '主机表', '工作表类型': '赋值表', '对称列数': 0, 'key列数': 1},
            {'工作簿名称': 'connection.xlsx', '工作表名称': '终端连接表', '工作表类型': '赋值表', '对称列数': 0, 'key列数': 2},
            {'工作簿名称': 'ipaddress.xlsx', '工作表名称': '网关地址表', '工作表类型': '赋值表', '对称列数': 0, 'key列数': 2},
            {'工作簿名称': 'parameter.xlsx', '工作表名称': '参数表', '工作表类型': '参数表', '对称列数': 0, 'key列数': 1},
        ])
        self._write_demo_excel(os.path.join(excel_dir, 'hostname.xlsx'), '主机表', [
            {'设备名': 'SW-ACCESS-01', '型号': 'H3C S5560X Demo', '角色': 'ASW', '楼层': '3F', '机柜': '弱电间A', 'U数': 36, '管理接口': 'Vlan-interface10', '管理IP': '192.168.10.11', '掩码': 24, 'SN': 'DEMO-SN-0001'},
        ])
        self._write_demo_excel(os.path.join(excel_dir, 'connection.xlsx'), '终端连接表', [
            {'己端设备': 'SW-ACCESS-01', '己端接口': 'GigabitEthernet1/0/1', '接入VLAN': 20, '接口类型': 'RJ45', '线缆类型': '网线', '终端名称': '办公区-PC-001', '备注信息': '办公网接入口'},
        ])
        self._write_demo_excel(os.path.join(excel_dir, 'ipaddress.xlsx'), '网关地址表', [
            {'己端设备': 'SW-ACCESS-01', '网关接口': 'Vlan-interface10', '管理VLAN': 10, '网关IP': '192.168.10.11', '网关掩码': 24, '备注': '管理地址'},
        ])
        self._write_demo_excel(os.path.join(excel_dir, 'parameter.xlsx'), '参数表', [
            {'全局参数名称': '本地用户名', '参数值': 'netadmin'},
            {'全局参数名称': '本地用户密钥', '参数值': 'ChangeMe_123'},
            {'全局参数名称': 'SSH使能', '参数值': 'yes'},
            {'全局参数名称': 'SSH端口', '参数值': 22},
            {'全局参数名称': '默认路由下一跳', '参数值': '192.168.10.1'},
            {'全局参数名称': 'NTP地址', '参数值': '192.168.10.100,192.168.10.101'},
            {'全局参数名称': 'LOGHOST地址', '参数值': '192.168.10.102'},
            {'全局参数名称': 'SNMP团体名', '参数值': 'demo_ro'},
            {'全局参数名称': 'SNMP地址', '参数值': '192.168.10.103'},
        ])

        asw_template = """# MagicCommander 示例：接入交换机开局配置
# 模板文件 ASW.j2 与 hostname.xlsx 中的 角色=ASW 对应
sysname {{ info['设备名'] }}
#
vlan {{ info['管理VLAN'] }}
 description MGMT
#
vlan {{ info['接入VLAN'] }}
 description OFFICE_ACCESS
#
interface {{ info['网关接口'] }}
 description {{ info['备注'] }}
 ip address {{ info['网关IP'] }} {{ info['网关掩码'] }}
#
interface {{ info['己端接口'] }}
 description TO-{{ info['终端名称'] }}
 port link-mode bridge
 port access vlan {{ info['接入VLAN'] }}
#
local-user {{ info['本地用户名'] }} class manage
 password simple {{ info['本地用户密钥'] }}
 service-type ssh terminal
 authorization-attribute user-role network-admin
#
{% if info['SSH使能'] == 'yes' %}
ssh server enable
ssh server port {{ info['SSH端口'] }}
{% endif %}
#
{% if info['NTP地址'][0] == 'list' %}
{% for ntp in info['NTP地址'][1:] %}
ntp-service unicast-server {{ ntp }}
{% endfor %}
{% endif %}
#
info-center loghost {{ info['LOGHOST地址'] }}
snmp-agent community read {{ info['SNMP团体名'] }}
snmp-agent target-host trap address udp-domain {{ info['SNMP地址'] }} params securityname {{ info['SNMP团体名'] }}
#
ip route-static 0.0.0.0 0.0.0.0 {{ info['默认路由下一跳'] }}
#
return
"""
        with open(os.path.join(templates_dir, 'ASW.j2'), 'w', encoding='utf-8') as fp:
            fp.write(asw_template)

        readme = """# 接入交换机配置示例

这是一个 MagicCommander 示例项目，用于演示如何通过 Excel + Jinja2 模板生成一台接入交换机的开局配置。

## 后端渲染规则

MagicCommander 当前按设备的“角色”字段选择模板：

- `excel/hostname.xlsx` 的 `角色` 字段为 `ASW`
- 渲染时会加载 `templates/ASW.j2`
- 模板中通过 `info['字段名']` 读取 Excel 汇总后的设备数据

因此，如果你把角色改成 `CORE`，就需要同步创建 `templates/CORE.j2`。

## 文件说明

- `para.xlsx`：声明后端要读取哪些 Excel、Sheet，以及读取类型。
- `excel/hostname.xlsx`：设备基础信息，包括设备名、角色、管理接口、管理 IP、SN。
- `excel/connection.xlsx`：示例接入口信息，包括接口名、终端名称、接入 VLAN。
- `excel/ipaddress.xlsx`：管理三层接口和管理 VLAN 信息。
- `excel/parameter.xlsx`：全局参数，如本地账号、SSH、NTP、Syslog、默认路由。
- `templates/ASW.j2`：接入交换机配置模板。

## 使用步骤

1. 在软件中选择该项目。
2. 查看或修改 `excel/*.xlsx` 中的示例参数。
3. 点击“渲染配置”，或在命令行执行：`python main.py render project <项目ID>`。
4. 到 `output/时间戳/ASW/` 查看生成的配置文件。
5. 到 `yaml/时间戳/ASW/` 查看中间 YAML 数据。

## 字段设计说明

当前后端会把同一台设备的多张表数据合并到一个 `info` 字典中。不同表里如果出现同名字段，后读取的值可能覆盖先读取的值。

所以本示例特意区分了：

- `接入VLAN`：用于接入口 `port access vlan`
- `管理VLAN`：用于管理 VLAN 和管理接口

请尽量避免在不同 Excel 表中重复使用含义不同的同名字段。

## 示例配置包含

- 设备名称
- 管理 VLAN
- 接入 VLAN
- 管理 IP
- 接入口描述
- 本地用户账号和密码
- SSH 服务
- NTP
- Syslog
- SNMP 简单示例
- 默认路由

## 安全提示

本示例账号和密码仅用于演示：

- 示例账号：`netadmin`
- 示例密码：`ChangeMe_123`

请勿在生产网络中直接使用示例密码。实际部署时，请按设备厂商要求使用加密口令、强密码策略、AAA/RADIUS/TACACS+ 等安全机制。

## 当前限制

第一版示例只放一条典型接入口。当前 `赋值表` 对同一台设备的多行普通字段不天然形成接口列表，后续如需批量生成多个接入口，建议进一步增强后端数据结构或使用可嵌套的表结构。
"""
        with open(os.path.join(project_dir, 'README.md'), 'w', encoding='utf-8') as fp:
            fp.write(readme)

    def execute_create(self, cmd_type: str, para: str, empty: bool = False):
        """实现项目创建功能"""
        mid_path = self._get_base_path()
        tar_path1 = os.path.join(mid_path, para)

        judge = False
        if not os.path.exists(tar_path1):
            os.makedirs(tar_path1, exist_ok=True)
        else:
            ans = input(f'{para} 项目目录已存在，请选择是否覆盖[y/n]')
            if ans.lower() == 'y':
                shutil.rmtree(tar_path1)
                judge = True
                os.makedirs(tar_path1, exist_ok=True)
            elif ans.lower() == 'n':
                return

        if not empty:
            self._create_demo_project_files(tar_path1)

        if judge:
            return

        mc_para_path = os.path.join(mid_path, 'MC_Para.xlsx')
        if os.path.exists(mc_para_path):
            df = pd.read_excel(mc_para_path)
            if para not in df['项目名称'].astype(str).tolist():
                # MC-S6: 仅追加 项目名称 列，保留其余列（勿整行覆盖为单列）
                df = df.copy()
                df.loc[len(df), '项目名称'] = para
                df.to_excel(mc_para_path, sheet_name='项目名称', index=False, header=True)
        else:
            pd.DataFrame({'项目名称': [para]}).to_excel(
                mc_para_path, sheet_name='项目名称', index=False, header=True
            )

    def execute_create_from_template(self, project_name: str, template_name: str):
        """从 example 目录复制模板项目并在 MC_Para.xlsx 中注册"""
        mid_path = self._get_base_path()
        example_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'example')
        src_dir = os.path.join(example_dir, template_name)
        tar_path = os.path.join(mid_path, project_name)

        if not os.path.exists(src_dir):
            raise FileNotFoundError(f'模板项目 "{template_name}" 不存在于 example 目录')

        if os.path.exists(tar_path):
            raise FileExistsError(f'项目 "{project_name}" 已存在')

        def _ignore_runtime_dirs(directory, contents):
            return [c for c in contents if c.startswith('.') or c == '__pycache__' or c in ('output', 'yaml', 'output-label')]

        shutil.copytree(src_dir, tar_path, ignore=_ignore_runtime_dirs)

        mc_para_path = os.path.join(mid_path, 'MC_Para.xlsx')
        if os.path.exists(mc_para_path):
            df = pd.read_excel(mc_para_path)
            if project_name not in df['项目名称'].astype(str).tolist():
                df.loc[len(df)] = [project_name]
                df.to_excel(mc_para_path, sheet_name='项目名称', index=False, header=True)
        else:
            pd.DataFrame({'项目名称': [project_name]}).to_excel(
                mc_para_path, sheet_name='项目名称', index=False, header=True
            )

    def execute_delete(self, cmd_type: str, para: str):
        """实现项目的目录删除功能"""
        mid_path = self._get_base_path()

        if cmd_type == 'project':
            if para == 'all':
                rem_name = list(self.project_name)
            else:
                rem_name = []
                for n in para.split('/'):
                    idx = int(n) - 1
                    if 0 <= idx < len(self.project_name):
                        rem_name.append(self.project_name[idx])

            for name in rem_name:
                dir_path = os.path.join(mid_path, name)
                self._safe_remove_dir(dir_path, f'{name} 项目删除成功', f'{name} 项目不存在，继续清理登记表')
                self._remove_from_mc_para(name)
                if name in self.project_name:
                    self.project_name.remove(name)
        else:
            # yaml, yaml-sn, output, output-sn, output-label
            folder_map = {
                'yaml': 'yaml',
                'yaml-sn': 'yaml-sn',
                'output': 'output',
                'output-sn': 'output-sn',
                'output-label': 'output-label',
            }
            folder_name = folder_map.get(cmd_type)
            if not folder_name:
                return

            if para == 'all':
                for name in self.project_name:
                    dir_path = os.path.join(mid_path, name, folder_name)
                    self._safe_remove_dir(dir_path, f'{name} 项目{folder_name}文件夹删除成功', f'{name} 不存在{folder_name}文件夹，删除无效')
            else:
                for n in para.split('/'):
                    name = self.project_name[int(n) - 1]
                    dir_path = os.path.join(mid_path, name, folder_name)
                    self._safe_remove_dir(dir_path, f'{name} 项目{folder_name}文件夹删除成功', f'{name} 不存在{folder_name}文件夹，删除无效')

    def _remove_from_mc_para(self, name: str):
        """从MC_Para表格中移除项目"""
        mid_path = self._get_base_path()
        mc_para_path = os.path.join(mid_path, 'MC_Para.xlsx')
        try:
            df = pd.read_excel(mc_para_path)
            df = df[~df['项目名称'].isin([name])]
            df.to_excel(mc_para_path, sheet_name='项目名称', index=False, header=True)
        except Exception as e:
            logger.error(f'更新MC_Para失败: {e}', exc_info=True)

    def execute_feature(self, cmd_type: str, para: str, config=None):
        """实现项目的标签卡转换 (config 支持纸张/方向/边距/每页数量/标签尺寸)"""
        mid_path = self._get_base_path()

        if cmd_type == 'label-print':
            time_str = strftime("%Y_%m_%d_%H_%M_%S", localtime())
            self.target_project_name = []
            self.target_project_num = []
            self.process_project_num(para)
            exceltolabel(self.target_project_name, time_str, config)
        elif cmd_type == 'label-md':
            time_str = strftime("%Y_%m_%d_%H_%M_%S", localtime())
            self.target_project_name = []
            self.target_project_num = []
            self.process_project_num(para)
            exceltomarkdown(self.target_project_name, time_str, config)
        elif cmd_type == 'label-delete':
            label_dir = 'output-label'
            if para == 'all':
                for name in self.project_name:
                    dir_path = os.path.join(mid_path, name, label_dir)
                    self._safe_remove_dir(dir_path, f'{name} 项目{label_dir}文件夹删除成功', f'{name} 不存在{label_dir}文件夹，删除无效')
            else:
                for n in para.split('/'):
                    name = self.project_name[int(n) - 1]
                    dir_path = os.path.join(mid_path, name, label_dir)
                    self._safe_remove_dir(dir_path, f'{name} 项目{label_dir}文件夹删除成功', f'{name} 不存在{label_dir}文件夹，删除无效')
