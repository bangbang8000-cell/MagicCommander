#!/usr/bin/env python3
"""
Magic Commander 3 - 命令行接口
网络设备配置管理工具 - 命令行版本
支持项目管理和配置渲染功能
"""

import argparse
import sys
import json
import os
import logging
import shutil
import tempfile
import zipfile
from pre_processing import PreProcessing
from config import WORKSPACE_DIR

logger = logging.getLogger(__name__)


def _project_origin(name: str) -> dict | None:
    """契约 v1.2（M-7）：从项目 template.meta.json 读取来源摘要（AL 项目 → MC 项目溯源）。"""
    meta_path = os.path.join(WORKSPACE_DIR, name, 'template.meta.json')
    if not os.path.exists(meta_path):
        return None
    try:
        with open(meta_path, 'r', encoding='utf-8') as f:
            m = json.load(f)
    except (OSError, ValueError):
        return None
    if not (m.get('originProjectId') or m.get('source')):
        return None
    return {
        'projectId': m.get('originProjectId', ''),
        'projectName': m.get('originProjectName', ''),
        'site': m.get('originSite', ''),
        'originPlan': m.get('originPlan', ''),
        'originPlanVersion': m.get('originPlanVersion'),
        'planHash': m.get('planHash', ''),
        'mcPlanVersion': m.get('mcPlanVersion'),
    }


def _load_plan_input(path: str) -> dict:
    """读取 plan 输入：.zip 交付包 → 解包取 plan.json；否则直接读 JSON（契约 v1.2 M-5）。"""
    if str(path).lower().endswith('.zip'):
        tmp = tempfile.mkdtemp(prefix='aidc_zip_')
        try:
            with zipfile.ZipFile(path) as z:
                _safe_extract_zip(z, tmp)
            plan_path = None
            for root, _, files in os.walk(tmp):
                for f in files:
                    if f == 'plan.json':
                        plan_path = os.path.join(root, f)
                        break
                if plan_path:
                    break
            if not plan_path:
                return {'error': f'交付包内未找到 plan.json: {path}'}
            with open(plan_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (zipfile.BadZipFile, ValueError) as e:
            # MC-S2: 恶意/损坏交付包整体拒绝并返回错误，不部分解压
            return {'error': f'交付包解压被拒绝: {e}'}
        finally:
            shutil.rmtree(tmp, ignore_errors=True)
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def _safe_extract_zip(zf: zipfile.ZipFile, dest_dir: str) -> None:
    """MC-S2（zip-slip 防护）：解压前逐条目校验，拒绝绝对路径 / 路径穿越条目。

    若任意条目不合法则整体拒绝（不部分解压），并抛出 ValueError。
    """
    dest_abs = os.path.abspath(dest_dir)
    for member in zf.infolist():
        name = member.filename.replace('\\', '/')
        # 拒绝绝对路径（盘符 / 前导 /）与 .. 穿越
        if name.startswith('/') or (len(name) > 1 and name[1] == ':'):
            raise ValueError(f'交付包条目含绝对路径: {member.filename}')
        parts = name.split('/')
        if '..' in parts:
            raise ValueError(f'交付包条目含路径穿越: {member.filename}')
    zf.extractall(dest_abs)


def _resolve_project_file(project_dir: str, rel_path: str) -> str:
    """安全拼接项目内相对路径，防止 '../' 穿越读取/写入项目目录之外的文件。"""
    project_abs = os.path.abspath(project_dir)
    full_path = os.path.abspath(os.path.join(project_abs, rel_path))
    if full_path != project_abs and not full_path.startswith(project_abs + os.sep):
        raise ValueError(f'文件路径不安全: {rel_path}')
    return full_path


def main():
    # 创建主解析器
    parser = argparse.ArgumentParser(
        description='Magic Commander 3 - 网络设备配置管理工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
        项目管理:
          mc project list                列出所有项目
          mc project create <name>       创建新项目
          mc project delete <id>         删除项目
          mc project info <id>          获取项目信息
        
        配置渲染:
          mc render project <ids>       渲染项目配置
          mc render yaml <ids>          渲染YAML文件
          mc render project-sn <ids>    渲染项目配置(SN模式)
          mc render yaml-sn <ids>       渲染YAML文件(SN模式)
        
        标签功能:
          mc label print <ids>          打印标签
          mc label delete <ids>         删除标签
        
        文件操作:
          mc file delete <type> <ids>   删除项目文件
          mc file list <id>             列出项目文件
        
        项目ID格式:
          - 单个ID: 1
          - 多个ID: 1,2,3
          - 所有项目: all
        
        示例:
          mc project list
          mc project create "test-project"
          mc render project 1
          mc render yaml 1,2,3
          mc render project-sn all
          mc label print 1
          mc file delete output 1
        '''
    )

    # 创建子解析器
    subparsers = parser.add_subparsers(title='命令类型', dest='command', help='可用命令')

    # AIDC 规划导入/分析（P1.4）
    plan_parser = subparsers.add_parser('plan', help='AIDC plan:table 导入与分析')
    plan_subparsers = plan_parser.add_subparsers(title='规划操作', dest='subcommand', help='规划子命令')
    plan_import_parser = plan_subparsers.add_parser('import', help='plan:table → MC 项目（契约 v1.2：按 projectId 自动匹配新建/更新/跳过；支持 .zip 交付包）')
    plan_import_parser.add_argument('plan_json', help='plan:table JSON 或 .zip 交付包路径')
    plan_import_parser.add_argument('project_dir', nargs='?', default=None,
                                    help='目标项目目录（缺省自动：命中按 projectId 更新，否则默认 projectName）')
    plan_import_parser.add_argument('--rehash', action='store_true',
                                    help='导入前按 macro 重算 planHash（GUI tunable 编辑路径，P2 V-MC4）')
    plan_analyze_parser = plan_subparsers.add_parser('analyze', help='j2 模板 ↔ 规划字段 对齐检查')
    plan_analyze_parser.add_argument('project_dir', help='项目目录')
    plan_validate_parser = plan_subparsers.add_parser('validate', help='专业校验（设备名/IP/AS/VLAN/网关）')
    plan_validate_parser.add_argument('plan_json', help='plan:table JSON 文件路径')
    plan_verify_parser = plan_subparsers.add_parser('verify', help='渲染命令核对矩阵（P2 V-MC2）')
    plan_verify_parser.add_argument('project_dir', help='项目目录')

    # 项目管理命令
    project_parser = subparsers.add_parser('project', help='项目管理操作')
    project_subparsers = project_parser.add_subparsers(title='项目操作', dest='subcommand', help='项目管理子命令')

    # 列出项目
    list_project_parser = project_subparsers.add_parser('list', help='列出所有项目')
    list_project_parser.add_argument('--format', choices=['text', 'json', 'yaml'], default='text', help='输出格式')

    # 创建项目
    create_project_parser = project_subparsers.add_parser('create', help='创建新项目')
    create_project_parser.add_argument('name', help='项目名称')
    create_project_parser.add_argument('--force', action='store_true', help='强制覆盖已存在的项目')
    create_project_parser.add_argument('--empty', action='store_true', help='创建空白项目（不生成示例文件）')
    create_project_parser.add_argument('--template', help='从 example 目录的模板创建项目')

    # 删除项目
    delete_project_parser = project_subparsers.add_parser('delete', help='删除项目')
    delete_project_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')
    delete_project_parser.add_argument('--force', action='store_true', help='强制删除项目，无需用户确认')

    # 项目信息
    info_project_parser = project_subparsers.add_parser('info', help='获取项目信息')
    info_project_parser.add_argument('id', help='项目ID')
    info_project_parser.add_argument('--format', choices=['text', 'json', 'yaml'], default='text', help='输出格式')

    # 4.8.0（F8-1 / 48-a）：项目包往返（导出可移植项目包 + 按身份导入）
    package_parser = project_subparsers.add_parser('package', help='项目包往返（导出/导入可移植项目包）')
    package_subparsers = package_parser.add_subparsers(title='项目包操作', dest='package_action', help='项目包子命令')
    package_export_parser = package_subparsers.add_parser('export', help='导出为可移植项目包（zip + manifest）')
    package_export_parser.add_argument('project_name', help='项目名称')
    package_export_parser.add_argument('output', help='目标 zip 路径')
    package_import_parser = package_subparsers.add_parser('import', help='导入项目包（按 manifest.projectId 匹配新建/更新/跳过）')
    package_import_parser.add_argument('package', help='项目包 zip 路径')
    package_import_parser.add_argument('project_dir', nargs='?', default=None,
                                       help='目标项目目录（缺省自动：命中按 projectId 更新，否则默认 projectName）')

    # 4.8.0（F8-3 / 48-c）：设备库可移植导入/导出（JSON/zip，schema+版本+条目清单）
    device_parser = subparsers.add_parser('device', help='设备库操作（4.8.0 F8-3 跨端资产互灌）')
    device_subparsers = device_parser.add_subparsers(title='设备库操作', dest='device_action', help='设备库子命令')
    device_lib_parser = device_subparsers.add_parser('library', help='设备库导入/导出')
    device_lib_sub = device_lib_parser.add_subparsers(title='设备库操作', dest='lib_action', help='设备库操作')
    device_export_parser = device_lib_sub.add_parser('export', help='导出设备库为可移植 JSON/zip')
    device_export_parser.add_argument('output', help='目标路径（.json 或 .zip）')
    device_import_parser = device_lib_sub.add_parser('import', help='导入设备库包（合并/去重/冲突提示）')
    device_import_parser.add_argument('package', help='设备库包路径（.json 或 .zip）')
    device_import_parser.add_argument('--target', default=None, help='目标 JSON（缺省写回内置 device_library.json）')

    # 渲染命令
    render_parser = subparsers.add_parser('render', help='配置渲染操作')
    render_subparsers = render_parser.add_subparsers(title='渲染操作', dest='subcommand', help='渲染子命令')

    # 渲染项目配置
    render_project_parser = render_subparsers.add_parser('project', help='渲染项目配置')
    render_project_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')
    render_project_parser.add_argument('--format', choices=['device_name', 'device_sn'], default='device_name', help='输出格式')

    # 渲染YAML文件
    render_yaml_parser = render_subparsers.add_parser('yaml', help='渲染YAML文件')
    render_yaml_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')
    render_yaml_parser.add_argument('--format', choices=['device_name', 'device_sn'], default='device_name', help='输出格式')

    # 渲染撤销
    render_undo_parser = render_subparsers.add_parser('undo', help='撤销渲染 (恢复最近一次备份)')
    render_undo_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')

    # 渲染预览
    render_dryrun_parser = render_subparsers.add_parser('dry-run', help='渲染预览 (不写文件，仅返回输出内容)')
    render_dryrun_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')
    render_dryrun_parser.add_argument('--format', choices=['device_name', 'device_sn'], default='device_name', help='输出格式')

    # 校验命令
    validate_parser = subparsers.add_parser('validate', help='校验操作')
    validate_subparsers = validate_parser.add_subparsers(title='校验操作', dest='subcommand', help='校验子命令')

    # 校验模板
    validate_template_parser = validate_subparsers.add_parser('template', help='校验 Jinja2 模板语法')
    validate_template_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')

    # 校验 Excel
    validate_excel_parser = validate_subparsers.add_parser('excel', help='校验 Excel 数据完整性')
    validate_excel_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')

    # 4.5.0（F5-1）：一致性校验引擎（参数表/模板/产物）
    validate_consistency_parser = validate_subparsers.add_parser('consistency', help='一致性校验（参数表完整性/模板映射/配置字段）')
    validate_consistency_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')
    # 4.5.0（F5-2）：导出数据核对（渲染批次 ↔ 参数/模板状态）
    validate_output_parser = validate_subparsers.add_parser('output', help='导出数据核对（数量/命名/引用/漂移）')
    validate_output_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')
    # 4.5.0（F5-3）：IP 规划校验（子网/网关/掩码/分配）
    validate_ip_parser = validate_subparsers.add_parser('ip', help='IP 规划校验（子网重叠/网关冲突/越界/重复/掩码）')
    validate_ip_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')
    # 4.5.0（F5-1~F5-3）：全量校验（一致性+导出核对+IP）
    validate_all_parser = validate_subparsers.add_parser('all', help='全量校验（一致性+导出核对+IP）')
    validate_all_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')
    # 4.8.0（F8-5 / 48-e）：交付物清单校验（缺失/漂移/哈希不符）
    validate_manifest_parser = validate_subparsers.add_parser('manifest', help='交付物清单校验（批次 manifest 缺失/漂移/哈希不符）')
    validate_manifest_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')

    # 4.8.0（F8-4 / 48-d）：项目评审报告/评审包（聚合校验报告 + 核对矩阵 + 项目摘要 + 交付清单）
    review_parser = subparsers.add_parser('review', help='项目评审报告/评审包（4.8.0 F8-4）')
    review_subparsers = review_parser.add_subparsers(title='评审操作', dest='review_action', help='评审子命令')
    review_report_parser = review_subparsers.add_parser('report', help='输出评审报告 JSON')
    review_report_parser.add_argument('project', help='项目名称')
    review_package_parser = review_subparsers.add_parser('package', help='导出评审包 zip（report.json + review.md + manifest）')
    review_package_parser.add_argument('project', help='项目名称')
    review_package_parser.add_argument('output', help='目标 zip 路径')
    review_md_parser = review_subparsers.add_parser('md', help='评审报告 → Markdown 文件（PDF 导出前置）')
    review_md_parser.add_argument('project', help='项目名称')
    review_md_parser.add_argument('output', help='目标 md 路径')

    # Diff 对比
    diff_parser = subparsers.add_parser('diff', help='对比渲染输出')
    diff_parser.add_argument('project', help='项目名称')
    diff_parser.add_argument('device', help='设备标识')
    diff_parser.add_argument('content', help='dry-run 渲染内容')
    diff_parser.add_argument('--format', choices=['device_name', 'device_sn'], default='device_name', help='输出格式')

    # 分析项目
    analyze_parser = subparsers.add_parser('analyze', help='分析项目模板和参数表')
    analyze_subparsers = analyze_parser.add_subparsers(title='分析操作', dest='subcommand', help='分析子命令')

    analyze_project_parser = analyze_subparsers.add_parser('project', help='分析项目')
    analyze_project_parser.add_argument('ids', help='项目ID或名称')

    # 智能校对
    proofread_parser = subparsers.add_parser('proofread', help='智能校对项目（模板语法/缺失列/数据空值）')
    proofread_subparsers = proofread_parser.add_subparsers(title='校对操作', dest='subcommand', help='校对子命令')

    proofread_project_parser = proofread_subparsers.add_parser('project', help='校对项目')
    proofread_project_parser.add_argument('ids', help='项目ID或名称')

    # 模板调试沙盒
    template_parser = subparsers.add_parser('template', help='模板操作')
    template_subparsers = template_parser.add_subparsers(title='模板操作', dest='subcommand', help='模板子命令')

    template_preview_parser = template_subparsers.add_parser('preview', help='调试沙盒：渲染指定模板文件')
    template_preview_parser.add_argument('ids', help='项目ID或名称')
    template_preview_parser.add_argument('template', help='模板文件相对路径（如 templates/ASW.j2）')

    # M6-b: 模板 CRUD（example 模板中心）
    template_list_parser = template_subparsers.add_parser('list', help='列出示例模板（example 目录）')
    template_list_parser.add_argument('--format', choices=['text', 'json'], default='json', help='输出格式')
    template_save_parser = template_subparsers.add_parser('save', help='将现有项目保存为示例模板')
    template_save_parser.add_argument('project', help='项目名称')
    template_save_parser.add_argument('name', help='新模板名称')
    template_update_parser = template_subparsers.add_parser('update', help='修改模板文件内容')
    template_update_parser.add_argument('name', help='模板名称')
    template_update_parser.add_argument('file_path', help='模板内文件相对路径（如 templates/ASW.j2）')
    template_update_parser.add_argument('content', help='新内容')
    template_delete_parser = template_subparsers.add_parser('delete', help='删除示例模板')
    template_delete_parser.add_argument('name', help='模板名称')
    template_delete_parser.add_argument('--force', action='store_true', help='无需确认直接删除')

    # 标签功能命令
    label_parser = subparsers.add_parser('label', help='标签功能操作')
    label_subparsers = label_parser.add_subparsers(title='标签操作', dest='subcommand', help='标签子命令')

    # 打印标签
    label_print_parser = label_subparsers.add_parser('print', help='打印标签')
    label_print_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')
    label_print_parser.add_argument('--config', help='JSON格式的打印配置 (纸张/方向/边距/每页数量/标签尺寸)', default=None)

    # 生成 Markdown 标签
    label_md_parser = label_subparsers.add_parser('md', help='生成Markdown标签')
    label_md_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')
    label_md_parser.add_argument('--config', help='JSON格式的标签配置', default=None)

    # 删除标签
    label_delete_parser = label_subparsers.add_parser('delete', help='删除标签')
    label_delete_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')

    # 文件操作命令
    file_parser = subparsers.add_parser('file', help='文件操作')
    file_subparsers = file_parser.add_subparsers(title='文件操作', dest='subcommand', help='文件操作子命令')

    # 删除项目文件
    file_delete_parser = file_subparsers.add_parser('delete', help='删除项目文件')
    file_delete_parser.add_argument('type', choices=['output', 'output-sn', 'yaml', 'yaml-sn'], help='删除文件类型')
    file_delete_parser.add_argument('ids', help='项目ID (使用,分隔多个ID)')
    file_delete_parser.add_argument('--force', action='store_true', help='强制删除文件，无需用户确认')

    # 列出项目文件
    file_list_parser = file_subparsers.add_parser('list', help='列出项目文件')
    file_list_parser.add_argument('id', help='项目ID')

    # 读取项目Excel文件
    read_excel_parser = project_subparsers.add_parser('read-excel', help='读取项目Excel文件')
    read_excel_parser.add_argument('id', help='项目ID')
    read_excel_parser.add_argument('file', help='Excel文件名')
    read_excel_parser.add_argument('--sheet', help='工作表名称（可选，默认第一个）')

    # 写入项目Excel文件
    write_excel_parser = project_subparsers.add_parser('write-excel', help='写入项目Excel文件')
    write_excel_parser.add_argument('id', help='项目ID')
    write_excel_parser.add_argument('file', help='Excel文件名')
    write_excel_parser.add_argument('data', help='JSON格式的写入数据')

    # 读取项目文本文件
    read_file_parser = project_subparsers.add_parser('read-file', help='读取项目文本文件')
    read_file_parser.add_argument('id', help='项目ID')
    read_file_parser.add_argument('path', help='相对于项目根目录的文件路径')

    # 写入项目文本文件
    write_file_parser = project_subparsers.add_parser('write-file', help='写入项目文本文件')
    write_file_parser.add_argument('id', help='项目ID')
    write_file_parser.add_argument('path', help='相对于项目根目录的文件路径')
    write_file_parser.add_argument('content', help='文件内容')

    # 列出项目文件（JSON格式）
    list_files_parser = project_subparsers.add_parser('list-files', help='列出项目文件（JSON格式）')
    list_files_parser.add_argument('id', help='项目ID')
    list_files_parser.add_argument('--type', choices=['excel', 'yaml', 'template', 'output', 'all'], default='all', help='文件类型过滤')

    # 全局选项
    parser.add_argument('--version', '-v', action='version', version='%(prog)s 3.0.0')
    parser.add_argument('--verbose', '-V', action='count', default=0, help='增加详细输出')
    parser.add_argument('--quiet', '-q', action='store_true', help='静默模式')

    args = parser.parse_args()

    # 如果没有提供命令，显示帮助信息
    if args.command is None:
        parser.print_help()
        sys.exit(0)

    try:
        processor = PreProcessing()
        processor.read_MC_para('MC_Para.xlsx')
        
        # 根据命令类型处理
        if args.command == 'project':
            handle_project_command(processor, args)
        elif args.command == 'plan':
            handle_plan_command(args)
        elif args.command == 'render':
            handle_render_command(processor, args)
        elif args.command == 'label':
            handle_label_command(processor, args)
        elif args.command == 'file':
            handle_file_command(processor, args)
        elif args.command == 'validate':
            handle_validate_command(processor, args)
        elif args.command == 'diff':
            handle_diff_command(processor, args)
        elif args.command == 'analyze':
            handle_analyze_command(processor, args)
        elif args.command == 'proofread':
            handle_proofread_command(processor, args)
        elif args.command == 'template':
            handle_template_command(processor, args)
        elif args.command == 'device':
            handle_device_command(args)
        elif args.command == 'review':
            handle_review_command(args)
        else:
            print_error(f'未知命令: {args.command}')
            sys.exit(1)

    except Exception as e:
        print_error(str(e))
        if args.verbose:
            logger.error("命令执行异常", exc_info=True)
        sys.exit(1)


def handle_plan_command(args):
    """AIDC plan:table 导入/分析（P1.4）。"""
    import json
    import os

    if args.subcommand == 'import':
        plan = _load_plan_input(args.plan_json)
        if 'error' in plan:
            print(f'plan 无效: {plan["error"]}')
            return
        # P2（V-MC4）：GUI tunable 编辑路径 —— 由 Python 权威按 macro 重算 planHash
        if args.rehash and plan.get('macro'):
            from intent.planner.validate import plan_hash
            plan['meta']['planHash'] = plan_hash(plan.get('macro', {}))
        from intent.planner.plantable_importer import import_plan_auto
        # 目标目录：缺省走自动（匹配/默认 projectName）；给定则作为显式目录（仍优先按 projectId 匹配更新）
        project_dir = args.project_dir
        if project_dir and not os.path.isabs(project_dir):
            project_dir = os.path.join(os.getcwd(), project_dir)
        summary = import_plan_auto(plan, WORKSPACE_DIR, explicit_dir=project_dir)
        if summary.get('error'):
            print(f'导入失败: {summary["error"]}')
            return
        # 注册到 MC_Para（可选：若在 workspace 下）
        proj_dir = summary.get('project_dir', '')
        mc_para = os.path.join(os.path.dirname(proj_dir), 'MC_Para.xlsx')
        if proj_dir and os.path.exists(mc_para):
            import pandas as pd
            df = pd.read_excel(mc_para)
            name = os.path.basename(proj_dir.rstrip('/'))
            if name not in df['项目名称'].astype(str).tolist():
                # MC-S6: 增量追加行，保留原表其余列（禁止重写为单列导致丢列）
                df = df.copy()
                df.loc[len(df), '项目名称'] = name
                df.to_excel(mc_para, sheet_name='项目名称', index=False)
        matched = summary.get('matched', 'none')
        if matched == 'skip':
            print(f'[skip] 规划无变化（v{summary.get("mcPlanVersion")}），跳过（{summary["name"]}）')
        else:
            verb = '更新' if matched == 'update' else ('新建' if matched == 'new' else '按目录导入')
            print(f'[OK] {verb} → {summary.get("name")}（v{summary.get("mcPlanVersion")}）')
            print(f'  设备 {summary.get("device_count")} / 接线 {summary.get("connections")} / 终端 {summary.get("terminals")}')
            for w in summary.get('warnings', []):
                print(f'  [warn] {w}')
            for c in summary.get('changelog', [])[-3:]:
                print(f'  [变更] {c.get("summary", "")}')
        if proj_dir and os.path.exists(mc_para):
            _df = pd.read_excel(mc_para)
            _names = _df['项目名称'].astype(str).tolist()
            if summary['name'] in _names:
                summary['mcpara_id'] = _names.index(summary['name']) + 1  # 1-based
        # 补充桥接标识（GUI 兼容）
        pmeta = plan.get('meta', {}) or {}
        summary['bridge'] = {k: pmeta.get(k) for k in ('source', 'projectType', 'bridgeVersion')}
        print(json.dumps(summary, ensure_ascii=False))
    elif args.subcommand == 'validate':
        from intent.planner.validate import validate_plan
        with open(args.plan_json, 'r', encoding='utf-8') as f:
            plan = json.load(f)
        issues = validate_plan(plan)
        print(f'[validate] plan:table 专业校验: {os.path.basename(args.plan_json)}')
        print(f'  问题数: {len(issues)}')
        for i in issues[:20]:
            print(f'    - {i}')
        print('  ' + ('PASS' if not issues else 'FAIL'))
        # G4：机器可读结果（GUI 校验按钮）
        print(json.dumps({'ok': not issues, 'issue_count': len(issues), 'issues': issues[:20]},
                         ensure_ascii=False))
    elif args.subcommand == 'analyze':
        from analyzer import analyze_project
        result = analyze_project(args.project_dir)
        missing = result.get('missing_columns', [])
        print(f'[analyze] 模板<->规划字段 对齐检查: {os.path.basename(args.project_dir)}')
        print(f'  缺失字段: {len(missing)}')
        for m in missing[:10]:
            print(f'    - {m}')
        print(f'  未被引用字段: {len(result.get("unused_columns", []))}')
        print(f'  复杂度: {result.get("complexity", 0)}')
    elif args.subcommand == 'verify':
        # P2（V-MC2）：渲染命令核对矩阵 → 结构化 JSON（GUI 命中矩阵可视化）
        import sys as _sys
        _sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'scripts'))
        from verify_rendered import verify_project_data
        result = verify_project_data(args.project_dir)
        print(json.dumps(result, ensure_ascii=False))


def _register_mc_para(proj_dir: str) -> None:
    """将项目注册到 workspace/MC_Para.xlsx（增量追加行，保留其余列）。"""
    mc_para = os.path.join(os.path.dirname(proj_dir), 'MC_Para.xlsx')
    if not proj_dir or not os.path.exists(mc_para):
        return
    import pandas as pd
    df = pd.read_excel(mc_para)
    name = os.path.basename(proj_dir.rstrip('/'))
    if name not in df['项目名称'].astype(str).tolist():
        df = df.copy()
        df.loc[len(df), '项目名称'] = name
        df.to_excel(mc_para, sheet_name='项目名称', index=False)


def handle_project_command(processor, args):
    """处理项目管理命令"""
    if args.subcommand == 'package':
        # 4.8.0（F8-1 / 48-a）：项目包往返（导出可移植项目包 + 按身份导入）
        from intent.planner.project_package import export_project_package, import_project_package
        if args.package_action == 'export':
            proj = os.path.join(WORKSPACE_DIR, args.project_name)
            if not os.path.isdir(proj):
                print_error(f'项目不存在: {args.project_name}')
                sys.exit(1)
            manifest = export_project_package(proj, args.output)
            summary = {'status': 'success', 'message': f'项目包已导出: {args.output}',
                       'data': {'path': args.output, 'projectId': manifest['projectId'],
                                'projectName': manifest['projectName'],
                                'file_count': manifest['summary']['file_count']}}
            print(json.dumps(summary, ensure_ascii=False))
            return
        if args.package_action == 'import':
            proj_dir = args.project_dir
            if proj_dir and not os.path.isabs(proj_dir):
                proj_dir = os.path.join(os.getcwd(), proj_dir)
            summary = import_project_package(args.package, WORKSPACE_DIR, explicit_dir=proj_dir)
            if summary.get('error'):
                print(f'导入失败: {summary["error"]}')
                return
            _register_mc_para(summary.get('project_dir', ''))
            matched = summary.get('matched', 'none')
            if matched == 'skip':
                print(f'[skip] 项目包无变化（{summary.get("name")}），跳过')
            else:
                verb = '更新' if matched == 'update' else '新建'
                print(f'[OK] {verb} → {summary.get("name")}（{summary.get("file_count")} 文件）')
            print(json.dumps(summary, ensure_ascii=False))
            return
    if args.subcommand == 'list':
        projects = []
        for i, name in enumerate(processor.project_name, 1):
            entry = {'id': i, 'name': name, 'index': i - 1}
            # 契约 v1.2（M-7）：来源摘要（AL 项目 → MC 项目溯源）
            origin = _project_origin(name)
            if origin:
                entry['origin'] = origin
            projects.append(entry)
        
        # 无论格式参数是什么，都返回 JSON 格式的输出
        print(json.dumps({
            'status': 'success',
            'message': '项目列表获取成功',
            'data': projects
        }, ensure_ascii=False, indent=2))

    elif args.subcommand == 'create':
        # 检查项目是否已存在
        if args.name in processor.project_name:
            if args.force:
                print_warning(f'项目 "{args.name}" 已存在，将强制覆盖')
            else:
                print_error(f'项目 "{args.name}" 已存在，请使用 --force 参数强制覆盖')
                sys.exit(1)
        
        if args.template:
            processor.execute_create_from_template(args.name, args.template)
        else:
            processor.execute_create('project', args.name, empty=args.empty)
        print_success(f'项目 "{args.name}" 创建成功')

    elif args.subcommand == 'delete':
        # 处理项目ID
        target_ids = process_project_ids(args.ids, processor.project_name)
        
        if not args.force:
            names = ', '.join([processor.project_name[idx] for idx in target_ids])
            confirm = input(f'确认删除项目: {names} [y/N]: ')
            if confirm.lower() != 'y':
                print_info('操作已取消')
                sys.exit(0)
        
        for idx in sorted(target_ids, reverse=True):
            project_name = processor.project_name[idx]
            processor.execute_delete('project', str(idx + 1))
            print_success(f'项目 "{project_name}" 删除成功')

    elif args.subcommand == 'info':
        try:
            project_id = int(args.id)
            if project_id < 1 or project_id > len(processor.project_name):
                raise ValueError(f'项目ID {args.id} 无效')
            project_name = processor.project_name[project_id - 1]
            
            info = {
                'id': project_id,
                'name': project_name,
                'path': os.path.join(WORKSPACE_DIR, project_name),
                'exists': os.path.exists(os.path.join(WORKSPACE_DIR, project_name))
            }
            
            if info['exists']:
                project_dir = info['path']
                info['structure'] = {
                    'excel': os.path.exists(os.path.join(project_dir, 'excel')),
                    'templates': os.path.exists(os.path.join(project_dir, 'templates')),
                    'para': os.path.exists(os.path.join(project_dir, 'para.xlsx')),
                    'output': os.path.exists(os.path.join(project_dir, 'output')),
                    'yaml': os.path.exists(os.path.join(project_dir, 'yaml'))
                }
            
            if args.format == 'json':
                print(json.dumps({
                    'status': 'success',
                    'message': '项目信息获取成功',
                    'data': info
                }, ensure_ascii=False, indent=2))
            elif args.format == 'yaml':
                try:
                    import yaml
                    print(yaml.dump({
                        'status': 'success',
                        'message': '项目信息获取成功',
                        'data': info
                    }, default_flow_style=False, allow_unicode=True))
                except ImportError:
                    print_error('YAML格式需要安装PyYAML库')
            else:
                logger.info(f'项目信息:')
                logger.info(f'ID: {info["id"]}')
                logger.info(f'名称: {info["name"]}')
                logger.info(f'路径: {info["path"]}')
                logger.info(f'存在: {"是" if info["exists"] else "否"}')
                if info['exists']:
                    logger.info('结构:')
                    for key, value in info['structure'].items():
                        logger.info(f'  - {key}: {"存在" if value else "不存在"}')
                        
        except ValueError as e:
            print_error(str(e))
            sys.exit(1)

    elif args.subcommand == 'read-excel':
        try:
            project_id = int(args.id)
            if project_id < 1 or project_id > len(processor.project_name):
                raise ValueError(f'项目ID {args.id} 无效')
            project_name = processor.project_name[project_id - 1]
            project_dir = os.path.join(WORKSPACE_DIR, project_name)
            
            file_path = _resolve_project_file(os.path.join(project_dir, 'excel'), args.file)
            if not os.path.exists(file_path):
                raise FileNotFoundError(f'文件不存在: {file_path}')

            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet_name = args.sheet or wb.sheetnames[0]
            if sheet_name not in wb.sheetnames:
                raise ValueError(f'工作表 "{sheet_name}" 不存在')
            
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else '' for cell in row])
            
            headers = rows[0] if rows else []
            data_rows = []
            for row in rows[1:]:
                obj = {}
                for i, header in enumerate(headers):
                    obj[header] = row[i] if i < len(row) else ''
                data_rows.append(obj)
            
            print(json.dumps({
                'status': 'success',
                'message': f'成功读取 {args.file} / {sheet_name}',
                'data': {
                    'name': sheet_name,
                    'headers': headers,
                    'rows': data_rows
                }
            }, ensure_ascii=False))
            
        except Exception as e:
            print(json.dumps({
                'status': 'error',
                'message': str(e),
                'data': None
            }, ensure_ascii=False))
            sys.exit(1)

    elif args.subcommand == 'write-excel':
        try:
            project_id = int(args.id)
            if project_id < 1 or project_id > len(processor.project_name):
                raise ValueError(f'项目ID {args.id} 无效')
            project_name = processor.project_name[project_id - 1]
            project_dir = os.path.join(WORKSPACE_DIR, project_name)
            
            file_path = _resolve_project_file(os.path.join(project_dir, 'excel'), args.file)
            if not os.path.exists(file_path):
                raise FileNotFoundError(f'文件不存在: {file_path}')

            data = json.loads(args.data)
            sheet_name = data.get('sheet', 'Sheet1')
            headers = data.get('headers', [])
            rows = data.get('rows', [])
            
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
            
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
            
            wb.save(file_path)
            print(json.dumps({
                'status': 'success',
                'message': f'成功写入 {args.file} / {sheet_name}',
                'data': None
            }, ensure_ascii=False))
            
        except Exception as e:
            print(json.dumps({
                'status': 'error',
                'message': str(e),
                'data': None
            }, ensure_ascii=False))
            sys.exit(1)

    elif args.subcommand == 'read-file':
        try:
            project_id = int(args.id)
            if project_id < 1 or project_id > len(processor.project_name):
                raise ValueError(f'项目ID {args.id} 无效')
            project_name = processor.project_name[project_id - 1]
            project_dir = os.path.join(WORKSPACE_DIR, project_name)
            
            file_path = _resolve_project_file(project_dir, args.path)
            if not os.path.exists(file_path):
                raise FileNotFoundError(f'文件不存在: {file_path}')

            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            print(json.dumps({
                'status': 'success',
                'message': f'成功读取 {args.path}',
                'data': content
            }, ensure_ascii=False))
            
        except Exception as e:
            print(json.dumps({
                'status': 'error',
                'message': str(e),
                'data': None
            }, ensure_ascii=False))
            sys.exit(1)

    elif args.subcommand == 'write-file':
        try:
            project_id = int(args.id)
            if project_id < 1 or project_id > len(processor.project_name):
                raise ValueError(f'项目ID {args.id} 无效')
            project_name = processor.project_name[project_id - 1]
            project_dir = os.path.join(WORKSPACE_DIR, project_name)
            
            file_path = _resolve_project_file(project_dir, args.path)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(args.content)
            
            print(json.dumps({
                'status': 'success',
                'message': f'成功写入 {args.path}',
                'data': None
            }, ensure_ascii=False))
            
        except Exception as e:
            print(json.dumps({
                'status': 'error',
                'message': str(e),
                'data': None
            }, ensure_ascii=False))
            sys.exit(1)

    elif args.subcommand == 'list-files':
        try:
            project_id = int(args.id)
            if project_id < 1 or project_id > len(processor.project_name):
                raise ValueError(f'项目ID {args.id} 无效')
            project_name = processor.project_name[project_id - 1]
            project_dir = os.path.join(WORKSPACE_DIR, project_name)
            
            if not os.path.exists(project_dir):
                raise FileNotFoundError(f'项目目录不存在: {project_dir}')
            
            def build_tree(dir_path, rel_path=''):
                entries = []
                try:
                    for entry in sorted(os.listdir(dir_path)):
                        if entry.startswith('.') or entry == '__pycache__':
                            continue
                        full = os.path.join(dir_path, entry)
                        rel = os.path.join(rel_path, entry) if rel_path else entry
                        if os.path.isdir(full):
                            children = build_tree(full, rel)
                            entries.append({
                                'name': entry,
                                'path': rel,
                                'isDirectory': True,
                                'children': children
                            })
                        else:
                            entries.append({
                                'name': entry,
                                'path': rel,
                                'isDirectory': False
                            })
                except PermissionError as e:
                    print(json.dumps({
                        'status': 'warning',
                        'message': f'跳过不可访问的目录: {e}',
                        'data': None
                    }, ensure_ascii=False))
                return entries
            
            files = build_tree(project_dir)
            
            print(json.dumps({
                'status': 'success',
                'message': f'项目 "{project_name}" 文件列表获取成功',
                'data': files
            }, ensure_ascii=False))
            
        except Exception as e:
            print(json.dumps({
                'status': 'error',
                'message': str(e),
                'data': None
            }, ensure_ascii=False))
            sys.exit(1)


def handle_render_command(processor, args):
    """处理配置渲染命令"""
    # 处理项目ID
    target_ids = process_project_ids(args.ids, processor.project_name)
    
    # 转换为项目编号字符串
    target_str = convert_to_project_string(target_ids)
    
    if args.subcommand == 'project':
        # 渲染项目配置
        format_type = 'device_sn' if args.format == 'device_sn' else 'device_name'
        processor.execute_render(target_str, format_type)
        print_success(f'项目配置渲染完成')
        
    elif args.subcommand == 'yaml':
        # 渲染YAML文件（device_sn 时输出到 yaml-sn 目录；不渲染配置文本）
        processor.execute_yaml(target_str, args.format)
        print_success(f'YAML文件渲染完成')

    elif args.subcommand == 'undo':
        # 撤销渲染：恢复最近一次备份
        restored = 0
        for project_id in target_ids:
            name = processor.project_name[project_id - 1]
            project_dir = os.path.join(WORKSPACE_DIR, name)
            if processor._restore_backup(project_dir):
                restored += 1
        print_success(f'已恢复 {restored} 个项目的渲染输出')

    elif args.subcommand == 'dry-run':
        # 渲染预览：不写文件，返回输出内容
        format_type = 'device_sn' if args.format == 'device_sn' else 'device_name'
        processor.execute_dry_run(target_str, format_type)


def handle_label_command(processor, args):
    """处理标签功能命令"""
    # 处理项目ID
    target_ids = process_project_ids(args.ids, processor.project_name)

    # 转换为项目编号字符串
    target_str = convert_to_project_string(target_ids)

    # 解析 --config 参数 (JSON)
    label_config = None
    if getattr(args, 'config', None):
        try:
            import json as _json
            label_config = _json.loads(args.config)
        except Exception as _e:
            logger.error(f"无法解析 config 参数: {_e}，使用默认配置")

    if args.subcommand == 'print':
        processor.execute_feature('label-print', target_str, label_config)
        print_success('标签打印完成')

    elif args.subcommand == 'md':
        processor.execute_feature('label-md', target_str, label_config)
        print_success('标签Markdown生成完成')

    elif args.subcommand == 'delete':
        processor.execute_feature('label-delete', target_str)
        print_success('标签删除完成')


def handle_validate_command(processor, args):
    """处理校验命令"""
    target_ids = process_project_ids(args.ids, processor.project_name)
    target_str = convert_to_project_string(target_ids)

    # 4.5.0（F5-1~F5-3）：校验引擎子命令 → 结构化报告（GUI/CI/测试复用）
    if args.subcommand in ('consistency', 'output', 'ip', 'all'):
        from validation import validate_project
        for idx in target_ids:
            name = processor.project_name[idx]
            project_dir = os.path.join(WORKSPACE_DIR, name)
            report = validate_project(project_dir, args.subcommand)
            print(report.to_json())
        return

    # 4.8.0（F8-5 / 48-e）：交付物清单校验（批次 manifest 缺失/漂移/哈希不符）
    if args.subcommand == 'manifest':
        from intent.delivery import verify_batch_manifest
        for idx in target_ids:
            name = processor.project_name[idx]
            project_dir = os.path.join(WORKSPACE_DIR, name)
            print(json.dumps(verify_batch_manifest(project_dir), ensure_ascii=False))
        return

    if args.subcommand == 'template':
        processor.validate_template(target_str)
    elif args.subcommand == 'excel':
        processor.validate_excel(target_str)

def handle_device_command(args):
    """设备库导入/导出（4.8.0 F8-3 跨端资产互灌）。"""
    from intent.device_library import export_device_library, import_device_library
    if args.device_action == 'library' and args.lib_action == 'export':
        bundle = export_device_library(args.output)
        summary = {'status': 'success', 'message': f'设备库已导出: {args.output}',
                   'data': {'path': args.output, 'schema': bundle['schema'],
                            'count': bundle['count']}}
        print(json.dumps(summary, ensure_ascii=False))
    elif args.device_action == 'library' and args.lib_action == 'import':
        result = import_device_library(args.package, target_path=args.target)
        print(json.dumps({'status': 'success',
                          'message': f"导入完成：新增 {len(result['added'])} / 更新 {len(result['updated'])} / 跳过 {len(result['skipped'])}",
                          'data': result}, ensure_ascii=False))
    else:
        print_error(f'未知设备库操作: {args.device_action}/{getattr(args, "lib_action", "")}')
        sys.exit(1)


def handle_review_command(args):
    """项目评审报告/评审包（4.8.0 F8-4）。"""
    from intent.review import build_review_package, build_review_report, write_review_markdown_file
    proj = os.path.join(WORKSPACE_DIR, args.project)
    if not os.path.isdir(proj):
        print_error(f'项目不存在: {args.project}')
        sys.exit(1)
    if args.review_action == 'report':
        report = build_review_report(proj)
        print(json.dumps(report, ensure_ascii=False))
    elif args.review_action == 'package':
        result = build_review_package(proj, args.output)
        print(json.dumps({'status': 'success', 'message': f'评审包已导出: {args.output}',
                          'data': {'path': result['path'], 'project': result['project']}},
                         ensure_ascii=False))
    elif args.review_action == 'md':
        result = write_review_markdown_file(proj, args.output)
        print(json.dumps({'status': 'success', 'message': f'评审 Markdown 已生成: {args.output}',
                          'data': {'path': result['path'], 'project': result['project']}},
                         ensure_ascii=False))
    else:
        print_error(f'未知评审操作: {getattr(args, "review_action", "")}')
        sys.exit(1)


def handle_diff_command(processor, args):
    """对比 dry-run 输出与已有输出文件"""
    import difflib
    project_dir = os.path.join(WORKSPACE_DIR, args.project)

    if args.format == 'device_sn':
        existing_path = os.path.join(project_dir, 'output-sn', f'conf_{args.device}.cfg')
    else:
        existing_path = os.path.join(project_dir, 'output', f'{args.device}.txt')

    existing_content = ''
    if os.path.exists(existing_path):
        try:
            with open(existing_path, 'r', encoding='utf-8') as f:
                existing_content = f.read()
        except:
            pass

    new_content = args.content
    if not existing_content:
        diff_lines = [f'[新增] 文件不存在: {existing_path}']
    else:
        differ = difflib.unified_diff(
            existing_content.splitlines(keepends=True),
            new_content.splitlines(keepends=True),
            fromfile=f'现有输出/{args.device}.txt',
            tofile=f'Dry-run 预览/{args.device}.txt',
            lineterm='',
        )
        diff_lines = list(differ)
        if not diff_lines:
            diff_lines = ['(无差异)']

    print(json.dumps({
        'status': 'success',
        'message': '对比完成',
        'data': {
            'diff': diff_lines,
            'hasExisting': bool(existing_content),
            'hasChanges': len(diff_lines) > 0 and diff_lines[0] != '(无差异)',
        },
    }, ensure_ascii=False))

def handle_analyze_command(processor, args):
    """分析项目模板和参数表"""
    from analyzer import analyze_project

    target_ids = process_project_ids(args.ids, processor.project_name)

    for idx in target_ids:
        project_name = processor.project_name[idx]
        project_path = os.path.join(WORKSPACE_DIR, project_name)

        report = analyze_project(project_path)
        print(json.dumps(report, ensure_ascii=False))


def handle_proofread_command(processor, args):
    """智能校对项目（模板语法/缺失列/数据空值）"""
    from proofread import proofread_project

    target_ids = process_project_ids(args.ids, processor.project_name)

    for idx in target_ids:
        project_name = processor.project_name[idx]
        project_path = os.path.join(WORKSPACE_DIR, project_name)

        report = proofread_project(project_path)
        print(json.dumps(report, ensure_ascii=False))


def _example_dir() -> str:
    """示例模板（模板中心）目录：backend 上级 / example"""
    return os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'example')


def _valid_template_name(name: str) -> bool:
    """模板名校验：非空、无路径分隔/穿越、非隐藏"""
    name = (name or '').strip()
    if not name or name.startswith(('.', '_')):
        return False
    if any(c in name for c in ('/', '\\', '..', ':', '*', '?', '"', '<', '>', '|')):
        return False
    return True


def handle_template_command(processor, args):
    """模板操作命令（M6-b：list/save/update/delete + preview）"""
    if args.subcommand == 'preview':
        # 安全校验模板相对路径，禁止穿越到项目外
        target_str = convert_to_project_string(process_project_ids(args.ids, processor.project_name))
        if '..' in args.template or args.template.startswith(('/', '\\')):
            print_error(f'模板路径无效: {args.template}')
            sys.exit(1)
        processor.execute_template_preview(target_str, args.template)
        print_success(f'模板预览完成')

    elif args.subcommand == 'list':
        ex = _example_dir()
        templates = []
        if os.path.isdir(ex):
            for entry in sorted(os.listdir(ex)):
                d = os.path.join(ex, entry)
                if os.path.isdir(d) and not entry.startswith(('.', '_')):
                    if os.path.exists(os.path.join(d, 'para.xlsx')) or os.path.isdir(os.path.join(d, 'templates')):
                        templates.append(entry)
        print(json.dumps({'status': 'success', 'message': '模板列表获取成功',
                          'data': templates}, ensure_ascii=False, indent=2))

    elif args.subcommand == 'save':
        if not _valid_template_name(args.name):
            print_error(f'模板名无效: {args.name}')
            sys.exit(1)
        proj = os.path.join(WORKSPACE_DIR, args.project)
        if not os.path.isdir(proj):
            print_error(f'项目不存在: {args.project}')
            sys.exit(1)
        target = os.path.join(_example_dir(), args.name)
        if os.path.exists(target):
            print_error(f'模板已存在: {args.name}')
            sys.exit(1)
        os.makedirs(_example_dir(), exist_ok=True)
        shutil.copytree(proj, target, ignore=shutil.ignore_patterns(
            'output', 'yaml', 'output-label', '.output_backups', '.render_cache'))
        print_success(f'模板 "{args.name}" 保存成功')

    elif args.subcommand == 'update':
        name = args.name
        if not _valid_template_name(name):
            print_error(f'模板名无效: {name}')
            sys.exit(1)
        if '..' in args.file_path or args.file_path.startswith(('/', '\\')):
            print_error(f'文件路径无效: {args.file_path}')
            sys.exit(1)
        target = os.path.join(_example_dir(), name, args.file_path)
        target_abs = os.path.abspath(target)
        root_abs = os.path.abspath(os.path.join(_example_dir(), name))
        if os.path.commonpath([target_abs, root_abs]) != root_abs:
            print_error('路径越界，禁止写入')
            sys.exit(1)
        if not os.path.isfile(target):
            print_error(f'文件不存在: {name}/{args.file_path}')
            sys.exit(1)
        os.makedirs(os.path.dirname(target), exist_ok=True)
        with open(target, 'w', encoding='utf-8') as f:
            f.write(args.content)
        print_success(f'模板文件更新: {name}/{args.file_path}')

    elif args.subcommand == 'delete':
        name = args.name
        if not _valid_template_name(name):
            print_error(f'模板名无效: {name}')
            sys.exit(1)
        target = os.path.join(_example_dir(), name)
        if not os.path.isdir(target):
            print_error(f'模板不存在: {name}')
            sys.exit(1)
        if not args.force:
            confirm = input(f'确认删除模板: {name} [y/N]: ')
            if confirm.lower() != 'y':
                print_info('操作已取消')
                sys.exit(0)
        shutil.rmtree(target, ignore_errors=True)
        print_success(f'模板 "{name}" 删除成功')

def handle_file_command(processor, args):
    """处理文件操作命令"""
    if args.subcommand == 'delete':
        # 删除项目文件
        target_ids = process_project_ids(args.ids, processor.project_name)
        
        if not args.force:
            file_type_name = {
                'output': '输出文件',
                'output-sn': 'SN模式输出文件',
                'yaml': 'YAML文件',
                'yaml-sn': 'SN模式YAML文件'
            }
            names = ', '.join([processor.project_name[idx] for idx in target_ids])
            confirm = input(f'确认删除{file_type_name.get(args.type, args.type)}: {names} [y/N]: ')
            if confirm.lower() != 'y':
                print_info('操作已取消')
                sys.exit(0)
        
        target_str = convert_to_project_string(target_ids)
        processor.execute_delete(args.type, target_str)
        print_success(f'{args.type} 文件删除完成')
        
    elif args.subcommand == 'list':
        # 列出项目文件
        try:
            project_id = int(args.id)
            if project_id < 1 or project_id > len(processor.project_name):
                raise ValueError(f'项目ID {args.id} 无效')
            
            project_name = processor.project_name[project_id - 1]
            project_dir = os.path.join(WORKSPACE_DIR, project_name)
            
            if not os.path.exists(project_dir):
                print_error(f'项目目录不存在: {project_dir}')
                sys.exit(1)
            
            logger.info(f'项目 "{project_name}" 文件结构:')
            for root, dirs, files in os.walk(project_dir):
                level = root.replace(project_dir, '').count(os.sep)
                indent = ' ' * 2 * level
                logger.info(f'{indent}{os.path.basename(root)}/')
                subindent = ' ' * 2 * (level + 1)
                for file in files:
                    logger.info(f'{subindent}{file}')
                    
        except ValueError as e:
            print_error(str(e))
            sys.exit(1)


def process_project_ids(ids_str, project_names):
    """处理项目ID字符串，支持数字编号和项目名称"""
    if ids_str.strip().lower() == 'all':
        return list(range(len(project_names)))
    
    try:
        ids = []
        for part in ids_str.strip().split(','):
            part = part.strip()
            if not part:
                continue
            if part.isdigit():
                # 数字编号
                idx = int(part) - 1
                if idx < 0 or idx >= len(project_names):
                    raise ValueError(f'项目ID {part} 无效，范围应在1-{len(project_names)}之间')
                ids.append(idx)
            else:
                # 项目名称
                if part in project_names:
                    ids.append(project_names.index(part))
                else:
                    raise ValueError(f'项目 "{part}" 不存在，可用项目: {", ".join(project_names)}')
        if not ids:
            raise ValueError(f'未指定有效的项目ID或名称')
        return ids
    except Exception as e:
        raise ValueError(f'无效的项目ID格式: {ids_str}')


def convert_to_project_string(ids):
    """将ID列表转换为项目编号字符串"""
    if len(ids) == 0:
        return ''
        
    if len(ids) == len(PreProcessing().project_name):
        return 'all'
        
    return '/'.join([str(idx + 1) for idx in ids])


def _safe_print(text: str):
    """安全打印，处理 Windows GBK 编码问题"""
    try:
        print(text)
    except UnicodeEncodeError:
        # GBK 编码回退：替换 Unicode 符号为 ASCII
        text = text.replace('\u2713', '[OK]').replace('\u2717', '[ERR]').replace('\u26a0', '[WARN]').replace('\u2139', '[INFO]')
        text = text.replace('\033[92m', '').replace('\033[91m', '').replace('\033[93m', '').replace('\033[94m', '').replace('\033[0m', '')
        print(text)


def print_success(message):
    """打印成功信息"""
    _safe_print(f'\033[92m✓ {message}\033[0m')


def print_error(message):
    """打印错误信息"""
    _safe_print(f'\033[91m✗ {message}\033[0m')


def print_warning(message):
    """打印警告信息"""
    _safe_print(f'\033[93m⚠ {message}\033[0m')


def print_info(message):
    """打印信息"""
    _safe_print(f'\033[94mℹ {message}\033[0m')


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print_error('\n操作被用户中断')
        sys.exit(1)
    except BrokenPipeError:
        # 处理管道中断错误
        sys.stderr.close()
        sys.exit(0)
