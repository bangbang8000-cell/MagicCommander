"""4.9.0（49-b / 49-d）：AIDC 示例资产生成与注册测试。

覆盖：
  - build_plan：设备/接线/终端数量、角色分布、命名规范、IB/RoCE 差异；
  - validate_plan：全部 plan:table 专业校验通过（含 planHash）；
  - 示例资产不变量：example/<key>/plan.json 与 build_plan 逐字段一致；
  - register_samples：落地结构（excel 四表/templates 8 角色/para/meta/README）与必填字段；
  - 导入回灌幂等：plan.json 导入 → 数据文件与示例逐字节一致。
"""
import importlib.util
import json
import os
import shutil
import tempfile

import pytest

from openpyxl import load_workbook

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SCRIPTS = os.path.join(REPO, 'scripts')
EXAMPLE_DIR = os.path.join(REPO, 'example')

SAMPLES_PY = os.path.join(SCRIPTS, 'aidc_samples.py')
VALIDATE_PY = os.path.join(SCRIPTS, 'validate_samples.py')


def _load_module(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


@pytest.fixture(scope='module')
def S():
    return _load_module('mc_aidc_samples', SAMPLES_PY)


@pytest.fixture(scope='module')
def V():
    return _load_module('mc_validate_samples', VALIDATE_PY)


class TestSampleDefs:
    def test_four_defs(self, S):
        assert [d['key'] for d in S.SAMPLE_DEFS] == ['64H100-IB', '64H100-RoCE', '128H100-IB', '128H100-RoCE']

    def test_scale_fields(self, S):
        for d in S.SAMPLE_DEFS:
            assert d['gpuCount'] in (64, 128)
            assert d['rails'] == 8
            assert d['leaves'] == 8
            assert d['spines'] == (2 if d['gpuCount'] == 64 else 4)
            assert d['fabric'] in ('ib', 'roce')
            assert d['project_id']
            assert d['description'] and d['scenario']


class TestBuildPlan:
    def test_device_counts(self, S):
        for d in S.SAMPLE_DEFS:
            plan = S.build_plan(d)
            assert S.device_count(plan) == (22 if d['gpuCount'] == 64 else 24)

    def test_role_distribution(self, S):
        expect = {
            'SPINE': 0, 'LEAF': 8, 'STO_SPINE': 1, 'STO_LEAF': 2,
            'BIZ_AGG': 2, 'BIZ_ACCESS': 4, 'OOB_AGG': 1, 'OOB_ACCESS': 2,
        }
        for d in S.SAMPLE_DEFS:
            plan = S.build_plan(d)
            roles = S.count_by_role(plan)
            assert roles['SPINE'] == d['spines']
            assert roles['LEAF'] == 8
            assert {k: v for k, v in roles.items() if k != 'SPINE'} == \
                   {k: v for k, v in expect.items() if k != 'SPINE'}

    def test_connections_terminals_counts(self, S):
        # 64 档：268 接线 / 720 终端；128 档：524 接线 / 1232 终端
        for d in S.SAMPLE_DEFS:
            plan = S.build_plan(d)
            if d['gpuCount'] == 64:
                assert len(plan['connections']) == 268
                assert len(plan['terminals']) == 720
            else:
                assert len(plan['connections']) == 524
                assert len(plan['terminals']) == 1232

    def test_naming_convention(self, S):
        for d in S.SAMPLE_DEFS:
            plan = S.build_plan(d)
            first = plan['deviceList'][0]
            site = d['site']
            vendor = d['vendor']
            assert first['name'].startswith(f'{site}-R01-AIDC-{vendor}-P-Spine-01')
            # 所有设备名规范：site-Rxx-AIDC-vendor-abbr-seq
            for dev in plan['deviceList']:
                assert dev['name'].startswith(f"{d['site']}-R{dev['rack']:02d}-AIDC-{d['vendor']}-")

    def test_ib_vs_roce_differences(self, S):
        plans = S.build_all_plans()
        ib = plans['64H100-IB']
        roce = plans['64H100-RoCE']
        assert ib['macro']['convergence'] == 1
        assert roce['macro']['convergence'] == 3
        assert ib['convergence']['compute'] == 1
        assert roce['convergence']['compute'] == 3
        # 型号矩阵差异（SPINE/LEAF/STO）
        assert ib['macro']['deviceModels']['SPINE'] != roce['macro']['deviceModels']['SPINE']
        assert ib['macro']['deviceModels']['LEAF'].startswith('NVIDIA')
        assert roce['macro']['deviceModels']['LEAF'].startswith('H3C')

    def test_plan_hash_matches_macro(self, S):
        from intent.planner.validate import plan_hash
        for d in S.SAMPLE_DEFS:
            plan = S.build_plan(d)
            assert plan['meta']['planHash'] == plan_hash(plan['macro'])

    def test_validate_plan_passes(self, S):
        from intent.planner.validate import validate_plan
        for d in S.SAMPLE_DEFS:
            plan = S.build_plan(d)
            assert validate_plan(plan) == []

    def test_meta_contract_v12(self, S):
        for d in S.SAMPLE_DEFS:
            plan = S.build_plan(d)
            meta = plan['meta']
            assert meta['schema'] == 'plan:table/1.2'
            assert meta['source'] == 'autolink'
            assert meta['projectType'] == 'aidc'
            assert meta['bridgeVersion'] == '1.0'
            assert meta['planVersion'] == 1
            assert meta['projectName'] == d['key']


class TestCommittedAssets:
    """示例资产不变量：example/ 下已注册资产与 build_plan 一致且结构完整。"""

    @pytest.mark.parametrize('key', ['64H100-IB', '64H100-RoCE', '128H100-IB', '128H100-RoCE'])
    def test_plan_json_matches_build_plan(self, S, key):
        defn = next(d for d in S.SAMPLE_DEFS if d['key'] == key)
        with open(os.path.join(EXAMPLE_DIR, key, 'plan.json'), encoding='utf-8') as f:
            committed = json.load(f)
        assert committed == S.build_plan(defn)

    @pytest.mark.parametrize('key', ['64H100-IB', '64H100-RoCE', '128H100-IB', '128H100-RoCE'])
    def test_structure_complete(self, S, key):
        base = os.path.join(EXAMPLE_DIR, key)
        assert os.path.isdir(base)
        for f in ('para.xlsx', 'plan.json', 'template.meta.json', 'README.md'):
            assert os.path.exists(os.path.join(base, f)), f'{key} 缺 {f}'
        for f in ('hostname.xlsx', 'connection.xlsx', 'ipaddress.xlsx', 'parameter.xlsx'):
            assert os.path.exists(os.path.join(base, 'excel', f)), f'{key} 缺 excel/{f}'
        j2 = [f for f in os.listdir(os.path.join(base, 'templates')) if f.endswith('.j2')]
        assert len(j2) == 8

    @pytest.mark.parametrize('key', ['64H100-IB', '64H100-RoCE', '128H100-IB', '128H100-RoCE'])
    def test_template_meta_required_fields(self, key):
        with open(os.path.join(EXAMPLE_DIR, key, 'template.meta.json'), encoding='utf-8') as f:
            meta = json.load(f)
        for k in ('name', 'description', 'scenario', 'inputRequirements', 'outputDescription'):
            assert meta.get(k), f'{key} template.meta 缺 {k}'
        # AIDC 桥接溯源保留
        assert meta['projectType'] == 'aidc'
        assert meta['originProjectId']

    @pytest.mark.parametrize('key', ['64H100-IB', '64H100-RoCE', '128H100-IB', '128H100-RoCE'])
    def test_validate_samples_load_passes(self, V, key):
        problems = []
        V.check_load(key, problems)
        assert problems == []


class TestRegister:
    def test_register_samples_structure(self, S):
        with tempfile.TemporaryDirectory() as tmp:
            example_dir = os.path.join(tmp, 'example')
            created = S.register_samples(example_dir=example_dir, workspace_dir=os.path.join(tmp, 'ws'))
            assert len(created) == 4
            for key, target in created:
                assert os.path.basename(target) == key
                assert os.path.exists(os.path.join(target, 'para.xlsx'))
                assert os.path.exists(os.path.join(target, 'plan.json'))
                assert os.path.exists(os.path.join(target, 'excel', 'hostname.xlsx'))
                assert os.path.exists(os.path.join(target, 'templates', 'SPINE.j2'))
                # 派生状态不入仓
                assert not os.path.exists(os.path.join(target, 'allocator_state.json'))

    def test_register_meta_normalized(self, S):
        with tempfile.TemporaryDirectory() as tmp:
            example_dir = os.path.join(tmp, 'example')
            S.register_samples(example_dir=example_dir, workspace_dir=os.path.join(tmp, 'ws'))
            with open(os.path.join(example_dir, '64H100-IB', 'template.meta.json'), encoding='utf-8') as f:
                meta = json.load(f)
            assert meta['updatedAt'] == S._FIXED_META_TS
            assert meta['description']
            assert meta['scenario']
            for entry in meta['changelog']:
                assert entry['at'] == S._FIXED_GENERATED_AT

    def test_roundtrip_plan_import_idempotent(self, S):
        """plan.json 导入新 workspace → 数据文件与示例逐字节一致。"""
        from intent.planner.plantable_importer import import_plan_auto
        with tempfile.TemporaryDirectory() as tmp:
            ws = os.path.join(tmp, 'ws')
            for key, _defn in ((d['key'], d) for d in S.SAMPLE_DEFS):
                plan_path = os.path.join(EXAMPLE_DIR, key, 'plan.json')
                with open(plan_path, encoding='utf-8') as f:
                    plan = json.load(f)
                r = import_plan_auto(plan, ws)
                assert r['ok']
                proj = r['project_dir']
                # 数据文件快照（排除 template.meta/README：示例为富化版）
                # xlsx 以逻辑内容（逐 sheet 单元格）比对，避免内嵌时间戳/压缩字节跨环境差异
                def data_snap(d):
                    out = {}
                    for root, dirs, files in os.walk(d):
                        dirs[:] = [x for x in dirs if x in ('excel', 'templates')]
                        for fname in files:
                            rel = os.path.relpath(os.path.join(root, fname), d).replace(os.sep, '/')
                            if fname in ('para.xlsx', 'plan.json') or rel.startswith(('excel/', 'templates/')):
                                p = os.path.join(root, fname)
                                if fname.endswith('.xlsx'):
                                    wb = load_workbook(p, read_only=True, data_only=False)
                                    sheets = {}
                                    for ws in wb.worksheets:
                                        rows = []
                                        for row in ws.iter_rows(values_only=True):
                                            rows.append([('' if v is None else v) for v in row])
                                        sheets[ws.title] = rows
                                    wb.close()
                                    out[rel] = sheets
                                elif fname == 'plan.json':
                                    with open(p, encoding='utf-8') as fh:
                                        out[rel] = json.load(fh)
                                else:
                                    with open(p, 'rb') as fh:
                                        out[rel] = fh.read()
                    return out
                assert data_snap(proj) == data_snap(os.path.join(EXAMPLE_DIR, key))


class TestValidateSamples:
    def test_expected_device_count(self, V):
        assert V.expected_device_count('64H100-IB') == 22
        assert V.expected_device_count('128H100-RoCE') == 24

    def test_render_snapshot_deterministic(self, V):
        fake = [
            {'device': 'A', 'role': 'SPINE', 'filename': 'A.txt', 'content': 'x'},
            {'device': 'B', 'role': 'LEAF', 'filename': 'B.txt', 'content': 'y'},
        ]
        s1 = V._render_snapshot(fake)
        s2 = V._render_snapshot(list(reversed(fake)))
        assert s1 == s2
        assert s1['device_count'] == 2
        assert s1['render_hash']

    def test_roundtrip_check_passes(self, V):
        problems = []
        V.check_roundtrip('64H100-IB', problems)
        assert problems == []

    def test_data_snapshot_excludes_runtime(self, V):
        with tempfile.TemporaryDirectory() as tmp:
            os.makedirs(os.path.join(tmp, 'excel'))
            os.makedirs(os.path.join(tmp, 'output'))
            with open(os.path.join(tmp, 'para.xlsx'), 'wb') as f:
                f.write(b'p')
            with open(os.path.join(tmp, 'output', 'x.txt'), 'wb') as f:
                f.write(b'x')
            snap = V._data_snapshot(tmp)
            assert 'para.xlsx' in snap
            assert 'output/x.txt' not in snap
