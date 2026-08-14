"""P1.2 专业表格（FR-E）测试：校验引擎 / Excel 格式 / 交叉引用。"""
import os
import tempfile

import pytest
from openpyxl import load_workbook

from intent.pilot64 import build_pilot64_context
from intent.project_single import generate_single_pilot64_project, SingleProjectGenerator
from intent.planner.validate import validate_context, validate_plan
from intent.planner.plantable import generate_plantable


class TestValidate:
    def test_valid_context(self):
        ctx = build_pilot64_context()
        issues = validate_context(ctx)
        assert issues == [], f'校验应通过，但: {issues}'

    def test_duplicate_hostname(self):
        ctx = build_pilot64_context()
        # 制造重复设备名
        ctx.device_params['LEAF'][2]['hostname_hostname_B_LEAF2'] = \
            ctx.device_params['LEAF'][1]['hostname_hostname_B_LEAF1']
        issues = validate_context(ctx)
        assert any('设备名重复' in i for i in issues)

    def test_invalid_ip(self):
        ctx = build_pilot64_context()
        ctx.device_params['LEAF'][1]['ipv4_M-ILO_P_LEAF1'] = '10.1.0.999/24'
        issues = validate_context(ctx)
        assert any('非法 IP' in i for i in issues)

    def test_queue_out_of_range(self):
        ctx = build_pilot64_context()
        ctx.globals['cnp_queue'] = 9
        issues = validate_context(ctx)
        assert any('cnp_queue' in i for i in issues)

    def test_validate_plan(self):
        ctx = build_pilot64_context()
        plan = generate_plantable(ctx)
        assert validate_plan(plan) == []


class TestProfessionalExcel:
    def test_formatting(self):
        ctx = build_pilot64_context()
        with tempfile.TemporaryDirectory() as tmp:
            project = os.path.join(tmp, 'aidc_pro')
            generate_single_pilot64_project(project, ctx)
            wb = load_workbook(os.path.join(project, 'excel', 'hostname.xlsx'))
            ws = wb['设备表-参数网']  # H2：设备表按四网拆 sheet
            assert ws.freeze_panes == 'A2'
            assert ws['A1'].font.bold is True
            # 多 sheet 工作簿（H2：四网拆 sheet）
            wb2 = load_workbook(os.path.join(project, 'excel', 'connection.xlsx'))
            assert {'终端连接表-参数网', '终端连接表-存储网', '终端连接表-业务网', '终端连接表-带外网',
                    'VLAN网关表-参数网', 'VLAN网关表-存储网', 'VLAN网关表-业务网'} <= set(wb2.sheetnames)
            wb3 = load_workbook(os.path.join(project, 'excel', 'ipaddress.xlsx'))
            assert {'IP规划地址表-参数网', 'IP规划地址表-存储网', 'IP规划地址表-业务网', 'IP规划地址表-带外网',
                    '环回地址表', '网段规划表'} <= set(wb3.sheetnames)

    def test_meta_validation(self):
        ctx = build_pilot64_context()
        with tempfile.TemporaryDirectory() as tmp:
            project = os.path.join(tmp, 'aidc_pro')
            generate_single_pilot64_project(project, ctx)
            import json
            meta = json.load(open(os.path.join(project, 'template.meta.json'), encoding='utf-8'))
            assert meta['validation']['ok'] is True
            assert meta['validation']['issue_count'] == 0


class TestCrossReference:
    def test_cross_reference(self):
        """设备↔接线↔IP↔VLAN 可追溯：接线中的设备均在设备表。"""
        ctx = build_pilot64_context()
        gen = SingleProjectGenerator(ctx)
        dev_names = set(gen.build_device_table()['设备名'])
        conn = gen.build_conn_table()
        for _i, row in conn.iterrows():
            assert row['己端设备'] in dev_names, f'接线引用未知设备: {row["己端设备"]}'
            assert row['对端设备'] in dev_names, f'接线引用未知对端: {row["对端设备"]}'
