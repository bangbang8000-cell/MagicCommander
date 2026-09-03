"""5.0.1（501-b）：validate_templates.py 参数合理性/协议兼容性检查测试。"""
import importlib.util
import json
import os
import tempfile

from openpyxl import Workbook

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
VT_PY = os.path.join(REPO, 'scripts', 'validate_templates.py')


def _load_vt():
    spec = importlib.util.spec_from_file_location('mc_validate_templates', VT_PY)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _write_plan(tpl, macro, device_list, connections):
    os.makedirs(tpl, exist_ok=True)
    with open(os.path.join(tpl, 'plan.json'), 'w', encoding='utf-8') as f:
        json.dump({'macro': macro, 'deviceList': device_list, 'connections': connections},
                  f, ensure_ascii=False)


def _write_param(tpl, rows):
    os.makedirs(os.path.join(tpl, 'excel'), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = '参数表'
    ws.append(['全局参数名', '参数值'])
    for k, v in rows:
        ws.append([k, v])
    wb.save(os.path.join(tpl, 'excel', 'parameter.xlsx'))


def _write_hostname(tpl, sheets):
    """sheets: {sheet名: [型号...]}，写 设备表-* 表（设备名/型号/角色）。"""
    os.makedirs(os.path.join(tpl, 'excel'), exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for name, models in sheets.items():
        ws = wb.create_sheet(name)
        ws.append(['设备名', '型号', '角色'])
        for i, m in enumerate(models, 1):
            ws.append([f'DEV-{i}', m, 'LEAF'])
    wb.save(os.path.join(tpl, 'excel', 'hostname.xlsx'))


def _base_macro(**overrides):
    macro = {
        'pfcQueue': 3, 'cnpQueue': 6, 'convergence': 3,
        'deviceModels': {'SPINE': 'H3C S9827', 'LEAF': 'H3C S9827',
                         'BIZ_AGG': 'H3C S9850-32H', 'BIZ_ACCESS': 'H3C S6850-56HF',
                         'OOB_AGG': 'H3C S6805-56HF-G', 'OOB_ACCESS': 'H3C S5560X-54C-EI',
                         'STO_SPINE': 'H3C S9825-128B', 'STO_LEAF': 'H3C S9825-128B'},
        'asRange': [65001, 65500],
        'vlanRanges': {'compute': [100, 199], 'storage': [200, 299],
                       'biz': [300, 399], 'oob': [400, 499]},
        'ipSegments': {'loopback': '10.1.0.0/20', 'compute': '10.1.16.0/20',
                       'storage': '10.1.32.0/20', 'biz': '10.1.48.0/20',
                       'oob': '10.1.64.0/21', 'interconnect': '10.1.72.0/21'},
    }
    macro.update(overrides)
    return macro


class TestParamReasonableness:
    def test_passes_on_reasonable_roce_plan(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        _write_plan(tpl, _base_macro(),
                    [{'name': 'X-R01-AIDC-H3C-P-Leaf-01', 'role': 'LEAF', 'asn': 65101}],
                    [{'src': 'X-R01-AIDC-H3C-P-Leaf-01', 'src_port': 'FourHundredGigE1/0/33', 'rate': '400G'}])
        _write_param(tpl, [('PFC队列', 3), ('CNP队列', 6)])
        with open(os.path.join(tpl, 'template.meta.json'), 'w', encoding='utf-8') as f:
            json.dump({'tunables': ['PFC队列', 'CNP队列']}, f)
        problems = []
        mod._check_param_reasonableness(tpl, problems)
        assert problems == []

    def test_pfc_queue_out_of_range(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        _write_plan(tpl, _base_macro(), [{'name': 'D-1', 'role': 'LEAF', 'asn': 65101}], [])
        _write_param(tpl, [('PFC队列', 9), ('CNP队列', 6)])
        problems = []
        mod._check_param_reasonableness(tpl, problems)
        assert any('PFC队列' in p and '0-7' in p for p in problems)

    def test_missing_tunable_key(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        _write_plan(tpl, _base_macro(), [{'name': 'D-1', 'role': 'LEAF', 'asn': 65101}], [])
        _write_param(tpl, [('PFC队列', 3)])
        with open(os.path.join(tpl, 'template.meta.json'), 'w', encoding='utf-8') as f:
            json.dump({'tunables': ['PFC队列', 'CNP队列']}, f)
        problems = []
        mod._check_param_reasonableness(tpl, problems)
        assert any('可调参数键' in p and 'CNP队列' in p for p in problems)

    def test_ib_convergence_must_be_one(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        macro = _base_macro(convergence=3, deviceModels={
            'SPINE': 'NVIDIA Quantum QM9700', 'LEAF': 'NVIDIA Quantum QM9700',
            'STO_SPINE': 'NVIDIA Quantum QM9700', 'STO_LEAF': 'NVIDIA Quantum QM9700',
            'BIZ_AGG': 'H3C S9850-32H', 'BIZ_ACCESS': 'H3C S6850-56HF',
            'OOB_AGG': 'H3C S6805-56HF-G', 'OOB_ACCESS': 'H3C S5560X-54C-EI'})
        _write_plan(tpl, macro, [{'name': 'D-1', 'role': 'LEAF', 'asn': 65101}], [])
        problems = []
        mod._check_param_reasonableness(tpl, problems)
        assert any('IB 收敛比' in p for p in problems)

    def test_roce_convergence_out_of_range(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        _write_plan(tpl, _base_macro(convergence=20), [{'name': 'D-1', 'role': 'LEAF', 'asn': 65101}], [])
        problems = []
        mod._check_param_reasonableness(tpl, problems)
        assert any('RoCE 收敛比' in p for p in problems)

    def test_bgp_as_out_of_range(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        _write_plan(tpl, _base_macro(), [{'name': 'D-1', 'role': 'LEAF', 'asn': 100}], [])
        problems = []
        mod._check_param_reasonableness(tpl, problems)
        assert any('AS' in p and '越出' in p for p in problems)

    def test_vlan_segment_overlap(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        macro = _base_macro(vlanRanges={'compute': [100, 199], 'storage': [150, 299]})
        _write_plan(tpl, macro, [{'name': 'D-1', 'role': 'LEAF', 'asn': 65101}], [])
        problems = []
        mod._check_param_reasonableness(tpl, problems)
        assert any('VLAN 段重叠' in p for p in problems)

    def test_ip_segment_overlap(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        macro = _base_macro(ipSegments={'loopback': '10.1.0.0/20', 'compute': '10.1.0.0/20'})
        _write_plan(tpl, macro, [{'name': 'D-1', 'role': 'LEAF', 'asn': 65101}], [])
        problems = []
        mod._check_param_reasonableness(tpl, problems)
        assert any('IP 段重叠' in p for p in problems)

    def test_connection_rate_exceeds_device_cap(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        # BIZ_ACCESS = S6850-56HF（多速率上限 100G）；400G 连接越限
        _write_plan(tpl, _base_macro(),
                    [{'name': 'ACC-1', 'role': 'BIZ_ACCESS', 'asn': 65141}],
                    [{'src': 'ACC-1', 'src_port': 'HundredGigE1/0/1', 'rate': '400G'}])
        problems = []
        mod._check_param_reasonableness(tpl, problems)
        assert any('连接速率越限' in p for p in problems)

    def test_legacy_no_plan_skips_plan_checks(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        os.makedirs(tpl, exist_ok=True)  # 无 plan.json
        _write_param(tpl, [('本地用户名', 'netadmin')])
        problems = []
        mod._check_param_reasonableness(tpl, problems)
        assert problems == []


class TestProtocolCompat:
    def test_passes_roce_models(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        _write_plan(tpl, _base_macro(), [{'name': 'D-1', 'role': 'LEAF', 'asn': 65101}], [])
        _write_hostname(tpl, {'设备表-参数网': ['H3C S9827'], '设备表-存储网': ['H3C S9825-128B']})
        problems = []
        mod._check_protocol_compat(tpl, problems)
        assert problems == []

    def test_plan_ib_requires_ib_switches(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        macro = _base_macro(deviceModels={
            'SPINE': 'H3C S9827', 'LEAF': 'H3C S9827',  # IB plan 却用 RoCE 交换机 → 报错
            'STO_SPINE': 'H3C S9825-128B', 'STO_LEAF': 'H3C S9825-128B',
            'BIZ_AGG': 'H3C S9850-32H', 'BIZ_ACCESS': 'H3C S6850-56HF',
            'OOB_AGG': 'H3C S6805-56HF-G', 'OOB_ACCESS': 'H3C S5560X-54C-EI'})
        # resolve_models_fabric 以 SPINE 是否 NVIDIA 判定 fabric；这里 SPINE=H3C → fabric=roce
        # 为构造「IB fabric 用 RoCE 交换机」，改为 NVIDIA SPINE + H3C LEAF
        macro['deviceModels'] = dict(macro['deviceModels'])
        macro['deviceModels']['SPINE'] = 'NVIDIA Quantum QM9700'
        _write_plan(tpl, macro, [{'name': 'D-1', 'role': 'LEAF', 'asn': 65101}], [])
        problems = []
        mod._check_protocol_compat(tpl, problems)
        assert any('协议不匹配' in p and 'LEAF' in p for p in problems)

    def test_unknown_plan_model(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        macro = _base_macro(deviceModels={
            'SPINE': 'H3C 不存在', 'LEAF': 'H3C S9827',
            'STO_SPINE': 'H3C S9825-128B', 'STO_LEAF': 'H3C S9825-128B',
            'BIZ_AGG': 'H3C S9850-32H', 'BIZ_ACCESS': 'H3C S6850-56HF',
            'OOB_AGG': 'H3C S6805-56HF-G', 'OOB_ACCESS': 'H3C S5560X-54C-EI'})
        _write_plan(tpl, macro, [{'name': 'D-1', 'role': 'LEAF', 'asn': 65101}], [])
        problems = []
        mod._check_protocol_compat(tpl, problems)
        assert any('不在设备库' in p for p in problems)

    def test_hostname_model_mismatch_with_plan_fabric(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        macro = _base_macro(deviceModels={
            'SPINE': 'NVIDIA Quantum QM9700', 'LEAF': 'NVIDIA Quantum QM9700',
            'STO_SPINE': 'NVIDIA Quantum QM9700', 'STO_LEAF': 'NVIDIA Quantum QM9700',
            'BIZ_AGG': 'H3C S9850-32H', 'BIZ_ACCESS': 'H3C S6850-56HF',
            'OOB_AGG': 'H3C S6805-56HF-G', 'OOB_ACCESS': 'H3C S5560X-54C-EI'})
        _write_plan(tpl, macro, [{'name': 'D-1', 'role': 'LEAF', 'asn': 65101}], [])
        # IB plan，但 excel 参数网型号为 H3C → 协议不一致
        _write_hostname(tpl, {'设备表-参数网': ['H3C S9827']})
        problems = []
        mod._check_protocol_compat(tpl, problems)
        assert any('协议' in p and '不一致' in p for p in problems)

    def test_legacy_no_plan_skipped(self, tmp_path):
        mod = _load_vt()
        tpl = str(tmp_path / 't')
        os.makedirs(tpl, exist_ok=True)
        problems = []
        mod._check_protocol_compat(tpl, problems)
        assert problems == []
